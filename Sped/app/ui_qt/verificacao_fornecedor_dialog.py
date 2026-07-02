"""
Verificação de Produtos por Fornecedor — verificação em lote.

Compara todos os produtos de todos os participantes/fornecedores encontrados
no SPED contra o catálogo da empresa. Mostra o que está e o que não está
cadastrado por fornecedor, com filtro e exportação.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path

from PySide6.QtCore import Qt, QThread, Signal, QObject
from PySide6.QtGui import QBrush, QColor, QPalette
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QRadioButton,
    QStyle,
    QStyledItemDelegate,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from app.repositories.mysql_cadastro import MysqlCadastroRepository

# ── Colunas ───────────────────────────────────────────────────────────────────

_MAIN_HEADERS = [
    "Status",
    "Fornecedor SPED",
    "Fornecedor Cadastro",
    "Código",
    "Descrição SPED",
    "NCM",
    "CST ICMS",
    "CFOP",
    "Alíq. ICMS %",
    "Período(s)",
    "Qtd. Docs",
]
_MAIN_FIELDS = [
    "status",
    "participant_name",
    "supplier_db_name",
    "code",
    "description",
    "ncm",
    "cst_icms",
    "cfop",
    "icms_rate",
    "periodos_str",
    "qtd_docs",
]

_EXPORT_EXTRA_HEADERS = [
    "Descrição Cadastro",
    "NCM Cadastro",
    "CST ICMS Cadastro",
    "CFOP Cadastro",
    "Alíq. ICMS Cadastro",
]
_EXPORT_EXTRA_FIELDS = [
    "descricao_cad",
    "ncm_cad",
    "cst_cad",
    "cfop_cad",
    "aliq_cad",
]

# Colunas da aba "Atualizar Fiscais via XML"
_FISCAL_HEADERS = [
    "✓", "Status", "Conf.", "Tipo Match", "Fornecedor", "Produto",
    "Cód. Empresa", "Cód. Forn.",
    "CFOP Saída Forn.", "Origem",
    "CST ICMS", "Alíq. ICMS %", "Red. BC ICMS %",
    "MVA %", "Vl. ICMS-ST", "Alíq. ICMS ST %",
    "CST PIS",     "Alíq. PIS %",
    "CST COFINS",  "Alíq. COFINS %",
    "Chave NFe",
]
_STATUS_COL = 1          # coluna "Status"
_FISCAL_FIELD_COLS = {
    # campo_db: índice na tabela (+1 em relação à versão anterior por causa do Status)
    "cfop_saida_fornecedor": 8,
    "origem_entrada":        9,
    "cst_icms":              10,
    "aliquota_icms":        11,
    "reducao_bc_icms":      12,
    "mva":                  13,
    "valor_icms_st":        14,
    "aliquota_icms_st":     15,
    "cst_pis":              16,
    "aliquota_pis":         17,
    "cst_cofins":           18,
    "aliquota_cofins":      19,
}
_FISCAL_KEY_COL = 20

# Status dos itens na aba fiscal
_FS_DIFF     = "Com diferença"
_FS_OK       = "Já correto"
_FS_NO_XML   = "Sem XML"
_FS_NO_CAD   = "Sem cadastro"
_FS_ALL      = [_FS_DIFF, _FS_OK, _FS_NO_XML, _FS_NO_CAD]
# Campos numéricos: DB e XML usam mesma unidade (percentual ou R$), sem conversão
_FISCAL_NUMERIC = frozenset({
    "aliquota_icms", "reducao_bc_icms", "mva",
    "valor_icms_st", "aliquota_icms_st",
    "aliquota_pis", "aliquota_cofins",
})

# ── Cores ─────────────────────────────────────────────────────────────────────

_BG_NAO   = QColor("#ffeb9c")   # amarelo — não cadastrado
_FG_NAO   = QColor("#7d4a00")
_BG_OK    = QColor("#c6efce")   # verde — já cadastrado
_FG_OK    = QColor("#276221")
_BG_SEM   = QColor("#d9e8f5")   # azul claro — sem correspondência no cadastro
_FG_SEM   = QColor("#1a4a7a")
_ROW_EVEN = QColor("#f3f6f9")
_ROW_ODD  = QColor("#ffffff")

_STATUS_NAO = "Não Cadastrado"
_STATUS_OK  = "Já Cadastrado"
_STATUS_SEM = "Sem Correspondência"


def _status_colors(status: str) -> tuple[QColor, QColor]:
    if status == _STATUS_OK:
        return _BG_OK, _FG_OK
    if status == _STATUS_SEM:
        return _BG_SEM, _FG_SEM
    return _BG_NAO, _FG_NAO


def _db_aliq_pct(v: object) -> str:
    """Converte fração de banco (0.18) para percentual display (18,00)."""
    s = str(v or "").strip()
    if not s:
        return ""
    try:
        result = round(Decimal(s.replace(",", ".")) * 100, 2)
        return f"{result:.2f}".replace(".", ",")
    except (InvalidOperation, Exception):
        return s


def _best_match_supplier(participant_name: str, suppliers: list[dict]) -> dict | None:
    """Retorna o fornecedor do cadastro com nome mais próximo do participante SPED."""
    pname_lower = participant_name.lower().strip()
    pname_words = set(pname_lower.split())
    best: dict | None = None
    best_score = 0
    for s in suppliers:
        sname = str(s.get("nome") or "").lower().strip()
        if not sname:
            continue
        if sname == pname_lower:
            return s
        swords = set(sname.split())
        score = len(pname_words & swords)
        if pname_lower in sname or sname in pname_lower:
            score += 5
        if score > best_score:
            best_score = score
            best = s
    return best if best_score >= 2 else None


# ── Delegate de pintura ───────────────────────────────────────────────────────

class _BrushDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index) -> None:
        painter.save()
        is_selected = bool(option.state & QStyle.State_Selected)
        bg = index.data(Qt.BackgroundRole)
        if is_selected:
            painter.fillRect(option.rect, option.palette.color(QPalette.Highlight))
        elif bg is not None:
            color = bg.color() if isinstance(bg, QBrush) else bg
            painter.fillRect(option.rect, color)
        if is_selected:
            painter.setPen(option.palette.color(QPalette.HighlightedText))
        else:
            fg = index.data(Qt.ForegroundRole)
            if fg is not None:
                painter.setPen(fg.color() if isinstance(fg, QBrush) else fg)
            else:
                painter.setPen(option.palette.color(QPalette.Text))
        font = index.data(Qt.FontRole)
        if font is not None:
            painter.setFont(font)
        painter.drawText(
            option.rect.adjusted(4, 0, -4, 0),
            Qt.AlignVCenter | Qt.AlignLeft,
            str(index.data(Qt.DisplayRole) or ""),
        )
        painter.restore()

    def sizeHint(self, option, index):
        hint = super().sizeHint(option, index)
        hint.setHeight(max(hint.height(), 22))
        return hint


# ── Worker: verificação em lote de todos os fornecedores ─────────────────────

class _VerificarTodosWorker(QObject):
    progress = Signal(int, int, str)
    finished = Signal(list)
    failed   = Signal(str)

    def __init__(
        self,
        rows: list[dict],
        operation_type: str,
        empresa_id: int,
        repository: MysqlCadastroRepository,
    ) -> None:
        super().__init__()
        self._rows           = rows
        self._operation_type = operation_type
        self._empresa_id     = empresa_id
        self._repo           = repository

    def run(self) -> None:
        try:
            # Campos de catálogo dependem da operação
            if self._operation_type == "Entrada":
                cst_f  = "cst_icms"
                cfop_f = "cfop_entrada"
                aliq_f = "aliquota_icms"
            else:
                cst_f  = "cst_icms_saida"
                cfop_f = "cfop_saida_empresa"
                aliq_f = "aliquota_icms_saida"

            # 1. Agrupa produtos do SPED por participante
            by_participant: dict[str, dict[str, dict]] = {}
            for row in self._rows:
                period = str(row.get("period") or "")
                for detail in row.get("launch_details", []):
                    participant = str(detail.get("participant_name") or "").strip()
                    if not participant:
                        continue
                    code = str(detail.get("code") or "").strip()
                    if not code:
                        continue
                    if participant not in by_participant:
                        by_participant[participant] = {}
                    prods = by_participant[participant]
                    if code not in prods:
                        prods[code] = {
                            "code":            code,
                            "description":     str(detail.get("description") or "").strip(),
                            "ncm":             str(detail.get("ncm") or "").strip(),
                            "cst_icms":        str(detail.get("cst_icms") or "").strip(),
                            "cfop":            str(detail.get("cfop") or "").strip(),
                            "icms_rate":       str(detail.get("icms_rate") or "").strip(),
                            "aliquota_pis":    str(detail.get("aliquota_pis") or "").strip(),
                            "aliquota_cofins": str(detail.get("aliquota_cofins") or "").strip(),
                            "periodos":        set(),
                            "qtd_docs":        0,
                            "document_keys":   [],
                        }
                    prod = prods[code]
                    if period:
                        prod["periodos"].add(period)
                    prod["qtd_docs"] += 1
                    if detail.get("document_key"):
                        key = str(detail["document_key"]).strip()
                        if key and key not in prod["document_keys"]:
                            prod["document_keys"].append(key)

            # 2. Carrega todos os fornecedores da empresa
            all_suppliers = self._repo.list_suppliers(self._empresa_id)

            # 3. Para cada participante, encontra o fornecedor e compara produtos
            all_rows: list[dict] = []
            participants = list(by_participant.keys())
            total = len(participants)

            catalog_cache: dict[int, tuple[dict, dict]] = {}

            for idx, participant in enumerate(participants):
                self.progress.emit(idx, total, f"Verificando: {participant[:50]}...")

                supplier = _best_match_supplier(participant, all_suppliers)

                sped_prods = by_participant[participant]

                if supplier is None:
                    # Sem correspondência no cadastro
                    for code, prod in sped_prods.items():
                        periodos = sorted(prod["periodos"])
                        all_rows.append({
                            **prod,
                            "participant_name":  participant,
                            "supplier_id":       None,
                            "supplier_db_name":  _STATUS_SEM,
                            "status":            _STATUS_SEM,
                            "periodos_str":      ", ".join(periodos),
                            "descricao_cad":     "",
                            "ncm_cad":           "",
                            "cst_cad":           "",
                            "cfop_cad":          "",
                            "aliq_cad":          "",
                        })
                    continue

                supplier_id = int(supplier["id"])
                supplier_name = str(supplier.get("nome") or "")

                # Carrega catálogo do fornecedor (com cache)
                if supplier_id not in catalog_cache:
                    catalog = self._repo.list_supplier_products(supplier_id)
                    cat_by_code: dict[str, dict] = {}
                    cat_by_ean:  dict[str, dict] = {}
                    for p in catalog:
                        for kf in ("codigo_empresa", "codigo_fornecedor"):
                            k = str(p.get(kf) or "").strip()
                            if k and k not in cat_by_code:
                                cat_by_code[k] = p
                        ean = "".join(c for c in str(p.get("ean") or "") if c.isdigit())
                        if ean and ean not in cat_by_ean:
                            cat_by_ean[ean] = p
                    catalog_cache[supplier_id] = (cat_by_code, cat_by_ean)
                else:
                    cat_by_code, cat_by_ean = catalog_cache[supplier_id]

                for code, prod in sped_prods.items():
                    cat = cat_by_code.get(code)
                    if cat is None:
                        digits = "".join(c for c in code if c.isdigit())
                        cat = cat_by_ean.get(digits) if digits else None

                    periodos = sorted(prod["periodos"])
                    base = {
                        **prod,
                        "participant_name": participant,
                        "supplier_id":      supplier_id,
                        "supplier_db_name": supplier_name,
                        "periodos_str":     ", ".join(periodos),
                    }

                    if cat is None:
                        base["status"]       = _STATUS_NAO
                        base["descricao_cad"] = ""
                        base["ncm_cad"]      = ""
                        base["cst_cad"]      = ""
                        base["cfop_cad"]     = ""
                        base["aliq_cad"]     = ""
                    else:
                        base["status"]        = _STATUS_OK
                        base["descricao_cad"] = str(cat.get("descricao") or "")
                        base["ncm_cad"]       = str(cat.get("ncm") or "")
                        base["cst_cad"]       = str(cat.get(cst_f) or "")
                        base["cfop_cad"]      = str(cat.get(cfop_f) or "")
                        base["aliq_cad"]      = _db_aliq_pct(cat.get(aliq_f))

                    all_rows.append(base)

            self.progress.emit(total, total, "Concluído.")
            self.finished.emit(all_rows)
        except Exception as exc:
            self.failed.emit(str(exc))


# ── Worker: cadastrar produtos para múltiplos fornecedores ───────────────────

class _CadastrarTodosWorker(QObject):
    progress = Signal(int, int, str)
    finished = Signal(dict)
    failed   = Signal(str)

    def __init__(
        self,
        produtos: list[dict],
        operation_type: str,
        repository: MysqlCadastroRepository,
    ) -> None:
        super().__init__()
        self._produtos       = produtos
        self._operation_type = operation_type
        self._repo           = repository

    def run(self) -> None:
        try:
            stats = {"cadastrados": 0, "ignorados": 0, "erros": []}
            total = len(self._produtos)

            for idx, prod in enumerate(self._produtos):
                supplier_id = prod.get("supplier_id")
                code        = str(prod.get("code") or "").strip()
                self.progress.emit(idx, total, f"Cadastrando {idx + 1}/{total} — {code}...")

                if not code or supplier_id is None:
                    stats["ignorados"] += 1
                    continue

                desc     = str(prod.get("description") or "").strip()
                ncm      = str(prod.get("ncm") or "").strip()
                cst      = str(prod.get("cst_icms") or "").strip()
                cfop     = str(prod.get("cfop") or "").strip()
                aliq_pct = str(prod.get("icms_rate") or "").strip()

                try:
                    aliq_frac = str(round(
                        Decimal(aliq_pct.replace(",", ".")) / 100, 4
                    )) if aliq_pct else "0"
                except InvalidOperation:
                    aliq_frac = "0"

                data: dict = {
                    "codigo_fornecedor": code,
                    "codigo_empresa":    code,
                    "ean":               code,
                    "descricao":         desc,
                    "ncm":              ncm,
                    "status_produto":    "ATIVO",
                }
                if self._operation_type == "Entrada":
                    data["cst_icms"]      = cst
                    data["cfop_entrada"]  = cfop
                    data["aliquota_icms"] = aliq_frac
                else:
                    data["cst_icms_saida"]     = cst
                    data["cfop_saida_empresa"]  = cfop
                    data["aliquota_icms_saida"] = aliq_frac

                try:
                    self._repo.save_supplier_product(supplier_id, data)
                    stats["cadastrados"] += 1
                except ValueError as exc:
                    if "Ja existe produto com o mesmo Codigo do Fornecedor" in str(exc):
                        existing = self._repo.fetch_supplier_product_by_fornecedor_code(
                            supplier_id, code
                        )
                        if existing and not str(existing.get("codigo_empresa") or "").strip():
                            data["id"] = existing["id"]
                            try:
                                self._repo.save_supplier_product(supplier_id, data)
                                stats["cadastrados"] += 1
                            except Exception as exc2:
                                stats["erros"].append({"code": code, "erro": str(exc2)})
                        else:
                            stats["erros"].append({
                                "code": code,
                                "erro": "Produto já existe para este fornecedor.",
                            })
                    else:
                        stats["erros"].append({"code": code, "erro": str(exc)})
                except Exception as exc:
                    stats["erros"].append({"code": code, "erro": str(exc)})

            self.progress.emit(total, total, "Concluído.")
            self.finished.emit(stats)
        except Exception as exc:
            self.failed.emit(str(exc))


# ── Diálogo principal ─────────────────────────────────────────────────────────

class VerificacaoFornecedorDialog(QDialog):
    """
    Verifica todos os fornecedores/participantes do SPED de uma só vez,
    mostrando quais produtos estão e quais não estão cadastrados no catálogo.
    Permite filtrar por fornecedor/status e exportar para Excel.
    """

    def __init__(
        self,
        rows: list[dict],
        operation_type: str,
        repository: MysqlCadastroRepository,
        environment: str,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._rows           = rows or []
        self._operation_type = operation_type
        self._repo           = repository
        self._env            = environment
        self._empresa_id: int | None = None
        self._all_rows: list[dict] = []
        self._filter_status    = "Todos"
        self._filter_supplier  = ""
        self._filter_btns: dict[str, QPushButton] = {}
        self._v_thread: QThread | None = None
        self._v_worker: _VerificarTodosWorker | None = None
        self._c_thread: QThread | None = None
        self._c_worker: _CadastrarTodosWorker | None = None
        # Tab 2 — Atualizar Fiscais via XML
        self._fiscal_proposals: list[dict] = []   # todos os itens
        self._f_displayed:      list[dict] = []   # subconjunto visível (após filtro)
        self._f_filter:         str        = "Todos"
        self._f_filter_btns:    dict       = {}
        self._f_thread:  QThread | None = None
        self._f_worker:  object | None = None
        self._fa_thread: QThread | None = None
        self._fa_worker: object | None = None
        # Tab 3 — Conferência XMLs × SPED
        self._conf_all_results: list[dict] = []
        self._conf_filter: str = "Todos"
        self._conf_filter_btns: dict[str, QPushButton] = {}
        self._conf_thread: QThread | None = None
        self._conf_worker: object | None = None

        participante = "Fornecedor" if operation_type == "Entrada" else "Cliente"

        self.setWindowTitle(f"Verificação de Produtos por {participante} — {operation_type}s")
        self.setMinimumSize(1200, 700)
        if parent:
            self.setStyleSheet(parent.styleSheet())
        self.resize(1450, 860)

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        self._tabs = QTabWidget()
        root.addWidget(self._tabs)

        tab1 = QWidget()
        self._build_tab_verificacao(tab1, participante)
        self._tabs.addTab(tab1, f"Verificação por {participante}")

        tab2 = QWidget()
        self._build_tab_fiscal_xml(tab2)
        self._tabs.addTab(tab2, "Atualizar Fiscais via XML")

        tab3 = QWidget()
        self._build_tab_conferencia_xml(tab3)
        self._tabs.addTab(tab3, "Conferência XMLs × SPED")

        # Fechar fica fora das abas
        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(12, 6, 12, 6)
        btn_row.addStretch()
        btn_fechar = QPushButton("Fechar")
        btn_fechar.clicked.connect(self.reject)
        btn_row.addWidget(btn_fechar)
        root.addLayout(btn_row)

    def _build_tab_verificacao(self, container: QWidget, participante: str) -> None:
        """Constrói o conteúdo da aba 1 — verificação de produtos por fornecedor."""
        lay = QVBoxLayout(container)
        lay.setContentsMargins(12, 10, 12, 10)
        lay.setSpacing(8)

        title = QLabel(f"Verificação de Produtos por {participante} — {self._operation_type}s")
        title.setObjectName("sectionTitle")
        lay.addWidget(title)

        # Empresa + verificar
        bar1 = QHBoxLayout()
        bar1.addWidget(QLabel("Empresa:"))
        self._company_combo = QComboBox()
        self._company_combo.setMinimumWidth(280)
        self._load_companies()
        bar1.addWidget(self._company_combo, 1)
        self._btn_verificar = QPushButton("Verificar Todos")
        self._btn_verificar.setObjectName("primaryButton")
        self._btn_verificar.clicked.connect(self._run_verificar)
        bar1.addWidget(self._btn_verificar)
        bar1.addStretch()
        lay.addLayout(bar1)

        # Métricas
        metrics_row = QHBoxLayout()
        self._metric_labels: dict[str, QLabel] = {}
        for key, label in (
            ("fornecedores", f"Fornecedores no SPED"),
            ("total",        "Total Produtos"),
            ("nao_cad",      "Não Cadastrados"),
            ("ja_cad",       "Já Cadastrados"),
            ("sem_corr",     "Sem Correspondência"),
        ):
            card = QWidget(); card.setObjectName("metricCard")
            cl = QVBoxLayout(card); cl.setContentsMargins(10, 6, 10, 6); cl.setSpacing(2)
            cl.addWidget(QLabel(label))
            val = QLabel("-"); val.setObjectName("metricValue"); cl.addWidget(val)
            self._metric_labels[key] = val
            metrics_row.addWidget(card)
        metrics_row.addStretch()
        lay.addLayout(metrics_row)

        self._status_label = QLabel(""); self._status_label.setObjectName("muted")
        lay.addWidget(self._status_label)

        # Filtros de status
        filter_bar = QHBoxLayout()
        filter_bar.addWidget(QLabel("Filtrar:"))
        _filter_styles = {
            "Todos":     "QPushButton:checked { background-color: #2563eb; color: white; font-weight: bold; }",
            _STATUS_NAO: "QPushButton:checked { background-color: #ffeb9c; color: #7d4a00; font-weight: bold; }",
            _STATUS_OK:  "QPushButton:checked { background-color: #c6efce; color: #276221; font-weight: bold; }",
            _STATUS_SEM: "QPushButton:checked { background-color: #d9e8f5; color: #1a4a7a; font-weight: bold; }",
        }
        for mode, css in _filter_styles.items():
            btn = QPushButton(mode); btn.setCheckable(True); btn.setChecked(mode == "Todos")
            btn.setStyleSheet(css)
            btn.clicked.connect(lambda _c, m=mode: self._set_status_filter(m))
            self._filter_btns[mode] = btn; filter_bar.addWidget(btn)
        filter_bar.addSpacing(20)
        filter_bar.addWidget(QLabel(f"{participante}:"))
        self._supplier_filter_combo = QComboBox(); self._supplier_filter_combo.setMinimumWidth(260)
        self._supplier_filter_combo.addItem(f"Todos os {participante}s", "")
        self._supplier_filter_combo.currentIndexChanged.connect(self._on_supplier_filter_changed)
        filter_bar.addWidget(self._supplier_filter_combo, 1); filter_bar.addStretch()
        lay.addLayout(filter_bar)

        # Tabela
        self._table = self._make_table(_MAIN_HEADERS)
        lay.addWidget(self._table, 1)

        # Barra inferior
        bottom = QHBoxLayout()
        self._btn_cadastrar = QPushButton("Cadastrar Não Cadastrados (visíveis)")
        self._btn_cadastrar.setObjectName("primaryButton"); self._btn_cadastrar.setEnabled(False)
        self._btn_cadastrar.clicked.connect(self._cadastrar); bottom.addWidget(self._btn_cadastrar)

        self._btn_export = QPushButton("Exportar Excel")
        self._btn_export.setEnabled(False); self._btn_export.clicked.connect(self._export)
        bottom.addWidget(self._btn_export)

        self._btn_atualizar_codigo = QPushButton("Atualizar Código Forn. (XML)")
        self._btn_atualizar_codigo.setEnabled(False)
        self._btn_atualizar_codigo.clicked.connect(self._abrir_atualizacao_codigo)
        bottom.addWidget(self._btn_atualizar_codigo)

        self._progress = QProgressBar(); self._progress.setRange(0, 100); self._progress.hide()
        bottom.addWidget(self._progress, 1)
        self._progress_label = QLabel(""); self._progress_label.hide()
        bottom.addWidget(self._progress_label)
        bottom.addStretch()
        lay.addLayout(bottom)

    def _build_tab_fiscal_xml(self, container: QWidget) -> None:
        """Constrói a aba 2 — atualização de campos fiscais via XMLs das NFe."""
        lay = QVBoxLayout(container)
        lay.setContentsMargins(12, 10, 12, 10)
        lay.setSpacing(8)

        title = QLabel("Atualizar Campos Fiscais do Catálogo via XML das NFe")
        title.setObjectName("sectionTitle")
        lay.addWidget(title)

        info = QLabel(
            "<b>Estratégias de vínculo produto ↔ XML (por ordem de prioridade):</b><br>"
            "1. <b>Por Código + Qtd/Val:</b> código e quantidade+valor coincidem na mesma NFe (máxima confiança)<br>"
            "2. <b>Por Código:</b> <code>codigo_fornecedor</code> = <code>cProd</code> na NFe do SPED<br>"
            "3. <b>Por Qtd/Val:</b> quantidade + valor ≈ item XML na NFe do SPED<br>"
            "4. <b>Por Cód. Forn.:</b> <code>codigo_fornecedor</code> = <code>cProd</code> em qualquer XML "
            "(apenas no escopo «Todo o cadastro»)<br>"
            "Apenas produtos com pelo menos um campo fiscal diferente do XML são exibidos."
        )
        info.setWordWrap(True); lay.addWidget(info)

        # Escopo de produtos
        scope_box = QGroupBox("Escopo de produtos:")
        scope_lay = QHBoxLayout(scope_box); scope_lay.setSpacing(20)
        self._f_radio_sped    = QRadioButton("Somente produtos do SPED carregado")
        self._f_radio_catalog = QRadioButton("Todo o cadastro de produtos")
        self._f_radio_sped.setChecked(True)
        self._f_scope_group = QButtonGroup(self)
        self._f_scope_group.addButton(self._f_radio_sped,    0)
        self._f_scope_group.addButton(self._f_radio_catalog, 1)
        scope_lay.addWidget(self._f_radio_sped)
        scope_lay.addWidget(self._f_radio_catalog)
        scope_lay.addStretch()
        lay.addWidget(scope_box)

        # Pasta de XMLs + scan
        folder_row = QHBoxLayout()
        folder_row.addWidget(QLabel("Pasta de XMLs:"))
        self._f_folder_edit = QLineEdit()
        self._f_folder_edit.setPlaceholderText("Selecione a pasta com os arquivos .xml das NFe...")
        self._f_folder_edit.setReadOnly(True)
        folder_row.addWidget(self._f_folder_edit, 1)
        btn_browse = QPushButton("Procurar...")
        btn_browse.clicked.connect(self._f_browse_folder)
        folder_row.addWidget(btn_browse)
        self._f_btn_scan = QPushButton("Escanear")
        self._f_btn_scan.setObjectName("primaryButton"); self._f_btn_scan.setEnabled(False)
        self._f_btn_scan.clicked.connect(self._f_run_scan)
        folder_row.addWidget(self._f_btn_scan)
        lay.addLayout(folder_row)

        # Seleção de campos fiscais a atualizar
        campos_box = QGroupBox("Campos a atualizar:")
        campos_lay = QHBoxLayout(campos_box); campos_lay.setSpacing(14)
        self._f_field_checks: dict[str, QCheckBox] = {}
        for key, lbl in (
            ("cfop_saida_fornecedor", "CFOP Saída Forn."),
            ("origem_entrada",        "Origem"),
            ("cst_icms",             "CST ICMS"),
            ("aliquota_icms",        "Alíq. ICMS %"),
            ("reducao_bc_icms",      "Red. BC ICMS %"),
            ("mva",                  "MVA %"),
            ("valor_icms_st",        "Vl. ICMS-ST"),
            ("aliquota_icms_st",     "Alíq. ICMS ST %"),
            ("cst_pis",              "CST PIS"),
            ("aliquota_pis",         "Alíq. PIS %"),
            ("cst_cofins",           "CST COFINS"),
            ("aliquota_cofins",      "Alíq. COFINS %"),
        ):
            chk = QCheckBox(lbl); chk.setChecked(True)
            chk.stateChanged.connect(self._f_refresh_table)
            self._f_field_checks[key] = chk; campos_lay.addWidget(chk)
        campos_lay.addStretch()
        lay.addWidget(campos_box)

        self._f_status_label = QLabel(""); self._f_status_label.setObjectName("muted")
        lay.addWidget(self._f_status_label)

        # Filtros por status
        filter_row = QHBoxLayout()
        for label in ["Todos", _FS_DIFF, _FS_OK, _FS_NO_XML, _FS_NO_CAD]:
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setChecked(label == "Todos")
            btn.clicked.connect(lambda checked, lbl=label: self._f_set_filter(lbl))
            self._f_filter_btns[label] = btn
            filter_row.addWidget(btn)
        filter_row.addStretch()
        lay.addLayout(filter_row)

        # Controles de linha
        sel_row = QHBoxLayout()
        for lbl, fn in (
            ("Selecionar Todos",   lambda: self._f_set_all(True)),
            ("Desmarcar Ambíguos", self._f_deselect_ambiguous),
            ("Desmarcar Todos",    lambda: self._f_set_all(False)),
        ):
            b = QPushButton(lbl); b.clicked.connect(fn); sel_row.addWidget(b)
        sel_row.addStretch()
        self._f_count_label = QLabel(""); sel_row.addWidget(self._f_count_label)
        lay.addLayout(sel_row)

        # Tabela
        self._f_table = QTableWidget()
        self._f_table.setColumnCount(len(_FISCAL_HEADERS))
        self._f_table.setHorizontalHeaderLabels(_FISCAL_HEADERS)
        self._f_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._f_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._f_table.verticalHeader().setVisible(False)
        self._f_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._f_table.horizontalHeader().setStretchLastSection(True)
        self._f_table.setAlternatingRowColors(False)
        self._f_table.setItemDelegate(_BrushDelegate(self._f_table))
        self._f_table.cellClicked.connect(self._f_on_cell_clicked)
        lay.addWidget(self._f_table, 1)

        self._f_progress = QProgressBar(); self._f_progress.hide()
        lay.addWidget(self._f_progress)

        # Barra inferior
        bot = QHBoxLayout()
        self._f_btn_apply = QPushButton("Aplicar Campos Selecionados")
        self._f_btn_apply.setObjectName("primaryButton"); self._f_btn_apply.setEnabled(False)
        self._f_btn_apply.clicked.connect(self._f_apply)
        bot.addWidget(self._f_btn_apply)
        self._f_btn_export = QPushButton("Exportar Excel")
        self._f_btn_export.setEnabled(False)
        self._f_btn_export.clicked.connect(self._f_export)
        bot.addWidget(self._f_btn_export)
        bot.addStretch()
        lay.addLayout(bot)

    def _build_tab_conferencia_xml(self, container: QWidget) -> None:
        """Aba 3 — confere se todos os XMLs das NFe do SPED estão presentes no disco."""
        lay = QVBoxLayout(container)
        lay.setContentsMargins(12, 10, 12, 10)
        lay.setSpacing(8)

        title = QLabel("Conferência de XMLs × Documentos do SPED")
        title.setObjectName("sectionTitle")
        lay.addWidget(title)

        info = QLabel(
            "Verifica se existe um arquivo XML para cada NFe registrada no SPED.<br>"
            "<b>Ausente no disco:</b> chave no SPED mas XML não encontrado na pasta.<br>"
            "<b>Encontrado:</b> chave no SPED e XML localizado na pasta.<br>"
            "<b>Extra no disco:</b> XML encontrado mas chave não está no SPED carregado."
        )
        info.setWordWrap(True)
        lay.addWidget(info)

        # Pasta + Escanear
        folder_row = QHBoxLayout()
        folder_row.addWidget(QLabel("Pasta de XMLs:"))
        self._conf_folder_edit = QLineEdit()
        self._conf_folder_edit.setPlaceholderText("Selecione a pasta com os arquivos .xml das NFe...")
        self._conf_folder_edit.setReadOnly(True)
        folder_row.addWidget(self._conf_folder_edit, 1)
        btn_browse = QPushButton("Procurar...")
        btn_browse.clicked.connect(self._conf_browse_folder)
        folder_row.addWidget(btn_browse)
        self._conf_btn_scan = QPushButton("Conferir")
        self._conf_btn_scan.setObjectName("primaryButton")
        self._conf_btn_scan.setEnabled(False)
        self._conf_btn_scan.clicked.connect(self._conf_run_scan)
        folder_row.addWidget(self._conf_btn_scan)
        lay.addLayout(folder_row)

        # Métricas
        metrics_row = QHBoxLayout()
        self._conf_metric_labels: dict[str, QLabel] = {}
        for key, lbl in (
            ("sped",       "Chaves no SPED"),
            ("xmls",       "XMLs no disco"),
            ("encontrado", "Encontrados"),
            ("ausente",    "Ausentes no disco"),
            ("extra",      "Extras no disco"),
        ):
            card = QWidget(); card.setObjectName("metricCard")
            cl = QVBoxLayout(card); cl.setContentsMargins(10, 6, 10, 6); cl.setSpacing(2)
            cl.addWidget(QLabel(lbl))
            val = QLabel("-"); val.setObjectName("metricValue"); cl.addWidget(val)
            self._conf_metric_labels[key] = val
            metrics_row.addWidget(card)
        metrics_row.addStretch()
        lay.addLayout(metrics_row)

        self._conf_status_label = QLabel(""); self._conf_status_label.setObjectName("muted")
        lay.addWidget(self._conf_status_label)

        # Filtros
        filter_bar = QHBoxLayout()
        filter_bar.addWidget(QLabel("Filtrar:"))
        _conf_styles = {
            "Todos":     "QPushButton:checked { background-color: #2563eb; color: white; font-weight: bold; }",
            "Ausente":   "QPushButton:checked { background-color: #f4cccc; color: #7d0000; font-weight: bold; }",
            "Encontrado":"QPushButton:checked { background-color: #c6efce; color: #276221; font-weight: bold; }",
            "Extra":     "QPushButton:checked { background-color: #d9e8f5; color: #1a4a7a; font-weight: bold; }",
        }
        for mode, css in _conf_styles.items():
            btn = QPushButton(mode); btn.setCheckable(True); btn.setChecked(mode == "Todos")
            btn.setStyleSheet(css)
            btn.clicked.connect(lambda _c, m=mode: self._conf_set_filter(m))
            self._conf_filter_btns[mode] = btn; filter_bar.addWidget(btn)
        filter_bar.addStretch()
        lay.addLayout(filter_bar)

        # Tabela
        _CONF_HEADERS = ["Status", "Número NF", "Data", "CNPJ Emitente", "Emitente",
                         "Valor Total", "Participante SPED", "Chave NFe", "Arquivo XML"]
        self._conf_table = QTableWidget()
        self._conf_table.setColumnCount(len(_CONF_HEADERS))
        self._conf_table.setHorizontalHeaderLabels(_CONF_HEADERS)
        self._conf_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._conf_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._conf_table.verticalHeader().setVisible(False)
        self._conf_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._conf_table.horizontalHeader().setStretchLastSection(True)
        self._conf_table.setItemDelegate(_BrushDelegate(self._conf_table))
        lay.addWidget(self._conf_table, 1)

        self._conf_progress = QProgressBar(); self._conf_progress.hide()
        lay.addWidget(self._conf_progress)

        # Barra inferior
        bot = QHBoxLayout()
        self._conf_btn_export = QPushButton("Exportar Excel")
        self._conf_btn_export.setEnabled(False)
        self._conf_btn_export.clicked.connect(self._conf_export)
        bot.addWidget(self._conf_btn_export)
        bot.addStretch()
        lay.addLayout(bot)

    # ── Helpers de UI ─────────────────────────────────────────────────────────

    def _make_table(self, headers: list[str]) -> QTableWidget:
        t = QTableWidget()
        t.setColumnCount(len(headers))
        t.setHorizontalHeaderLabels(headers)
        t.setAlternatingRowColors(False)
        t.setSelectionBehavior(QAbstractItemView.SelectRows)
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.verticalHeader().setVisible(False)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        t.horizontalHeader().setStretchLastSection(True)
        t.setItemDelegate(_BrushDelegate(t))
        return t

    def _load_companies(self) -> None:
        try:
            companies = self._repo.list_companies_for_confronto(self._env)
            for c in companies:
                label = str(c["nome"] or "").strip() or str(c["cnpj"] or "")
                self._company_combo.addItem(label, str(c["cnpj"] or ""))
            if len(companies) == 1:
                self._company_combo.setCurrentIndex(0)
        except Exception:
            pass

    def _resolve_empresa_id(self) -> int | None:
        cnpj = str(self._company_combo.currentData() or "").strip()
        name = self._company_combo.currentText()
        if not cnpj:
            return None
        try:
            return self._repo.find_company_id(self._env, name, cnpj)
        except Exception:
            return None

    # ── Verificação em lote ───────────────────────────────────────────────────

    def _run_verificar(self) -> None:
        if not self._rows:
            QMessageBox.warning(self, "Verificar", "Não há dados no SPED. Processe o SPED primeiro.")
            return

        empresa_id = self._resolve_empresa_id()
        if empresa_id is None:
            QMessageBox.warning(self, "Verificar", "Selecione uma empresa.")
            return

        self._empresa_id = empresa_id
        self._btn_verificar.setEnabled(False)
        self._btn_cadastrar.setEnabled(False)
        self._btn_export.setEnabled(False)
        self._progress.setValue(0)
        self._progress.show()
        self._progress_label.setText("Iniciando...")
        self._progress_label.show()
        self._status_label.setText("")

        self._v_thread = QThread(self)
        self._v_worker = _VerificarTodosWorker(
            self._rows, self._operation_type, empresa_id, self._repo
        )
        self._v_worker.moveToThread(self._v_thread)
        self._v_thread.started.connect(self._v_worker.run)
        self._v_worker.progress.connect(self._on_v_progress)
        self._v_worker.finished.connect(self._on_v_done)
        self._v_worker.failed.connect(self._on_v_failed)
        self._v_worker.finished.connect(self._v_thread.quit)
        self._v_worker.failed.connect(self._v_thread.quit)
        self._v_thread.finished.connect(self._v_worker.deleteLater)
        self._v_thread.finished.connect(self._v_thread.deleteLater)
        self._v_thread.finished.connect(lambda: setattr(self, "_v_thread", None))
        self._v_thread.finished.connect(lambda: setattr(self, "_v_worker", None))
        self._v_thread.start()

    def _on_v_progress(self, current: int, total: int, msg: str) -> None:
        if total > 0:
            self._progress.setValue(int(current / total * 100))
        self._progress_label.setText(msg)

    def _on_v_done(self, all_rows: list[dict]) -> None:
        self._progress.hide()
        self._progress_label.hide()
        self._btn_verificar.setEnabled(True)
        self._all_rows = all_rows

        # Atualiza combo de filtro por fornecedor
        self._supplier_filter_combo.blockSignals(True)
        self._supplier_filter_combo.clear()
        participante = "Fornecedor" if self._operation_type == "Entrada" else "Cliente"
        self._supplier_filter_combo.addItem(f"Todos os {participante}s", "")
        participants = sorted({r["participant_name"] for r in all_rows}, key=str.lower)
        for p in participants:
            self._supplier_filter_combo.addItem(p, p)
        self._supplier_filter_combo.blockSignals(False)

        self._filter_status   = "Todos"
        self._filter_supplier = ""
        for m, btn in self._filter_btns.items():
            btn.setChecked(m == "Todos")
        self._supplier_filter_combo.setCurrentIndex(0)

        self._fill_table(all_rows)
        self._update_metrics(all_rows)
        self._update_filter_btns(all_rows)
        self._apply_filter()

        nao = sum(1 for r in all_rows if r["status"] == _STATUS_NAO)
        ja  = sum(1 for r in all_rows if r["status"] == _STATUS_OK)
        sem = sum(1 for r in all_rows if r["status"] == _STATUS_SEM)
        self._status_label.setText(
            f"Verificação concluída: {len(all_rows)} produto(s) — "
            f"{nao} não cadastrado(s) · {ja} já cadastrado(s) · {sem} sem correspondência"
        )
        self._btn_export.setEnabled(bool(all_rows))
        self._btn_atualizar_codigo.setEnabled(bool(all_rows))
        self._btn_cadastrar.setEnabled(nao > 0)

    def _on_v_failed(self, msg: str) -> None:
        self._progress.hide()
        self._progress_label.hide()
        self._btn_verificar.setEnabled(True)
        QMessageBox.critical(self, "Verificar", f"Erro ao verificar:\n{msg}")

    # ── Preenchimento da tabela ───────────────────────────────────────────────

    def _fill_table(self, rows: list[dict]) -> None:
        self._table.setUpdatesEnabled(False)
        self._table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            status = str(row.get("status") or "")
            bg_status, fg_status = _status_colors(status)
            row_bg = _ROW_EVEN if r % 2 == 0 else _ROW_ODD
            for c, field in enumerate(_MAIN_FIELDS):
                val = row.get(field)
                text = str(val if val is not None else "")
                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                if c == 0:
                    item.setBackground(bg_status)
                    item.setForeground(fg_status)
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                else:
                    item.setBackground(row_bg)
                self._table.setItem(r, c, item)
        self._table.resizeColumnsToContents()
        self._table.setUpdatesEnabled(True)

    # ── Filtros ───────────────────────────────────────────────────────────────

    def _row_visible(self, row: dict) -> bool:
        if self._filter_status != "Todos" and row.get("status") != self._filter_status:
            return False
        if self._filter_supplier and row.get("participant_name") != self._filter_supplier:
            return False
        return True

    def _apply_filter(self) -> None:
        self._table.setUpdatesEnabled(False)
        visible_nao = 0
        for r, row in enumerate(self._all_rows):
            visible = self._row_visible(row)
            self._table.setRowHidden(r, not visible)
            if visible and row.get("status") == _STATUS_NAO:
                visible_nao += 1
        self._table.setUpdatesEnabled(True)
        self._btn_cadastrar.setEnabled(visible_nao > 0)

    def _set_status_filter(self, mode: str) -> None:
        self._filter_status = mode
        for m, btn in self._filter_btns.items():
            btn.setChecked(m == mode)
        self._apply_filter()

    def _on_supplier_filter_changed(self) -> None:
        self._filter_supplier = str(self._supplier_filter_combo.currentData() or "")
        self._apply_filter()

    def _update_filter_btns(self, rows: list[dict]) -> None:
        total = len(rows)
        nao   = sum(1 for r in rows if r["status"] == _STATUS_NAO)
        ja    = sum(1 for r in rows if r["status"] == _STATUS_OK)
        sem   = sum(1 for r in rows if r["status"] == _STATUS_SEM)
        self._filter_btns["Todos"].setText(f"Todos ({total})")
        self._filter_btns[_STATUS_NAO].setText(f"Não Cadastrados ({nao})")
        self._filter_btns[_STATUS_OK].setText(f"Já Cadastrados ({ja})")
        self._filter_btns[_STATUS_SEM].setText(f"Sem Correspondência ({sem})")

    def _update_metrics(self, rows: list[dict]) -> None:
        fornecedores = len({r["participant_name"] for r in rows})
        nao  = sum(1 for r in rows if r["status"] == _STATUS_NAO)
        ja   = sum(1 for r in rows if r["status"] == _STATUS_OK)
        sem  = sum(1 for r in rows if r["status"] == _STATUS_SEM)
        self._metric_labels["fornecedores"].setText(str(fornecedores))
        self._metric_labels["total"].setText(str(len(rows)))
        self._metric_labels["nao_cad"].setText(str(nao))
        self._metric_labels["ja_cad"].setText(str(ja))
        self._metric_labels["sem_corr"].setText(str(sem))

    # ── Cadastrar ─────────────────────────────────────────────────────────────

    def _cadastrar(self) -> None:
        # Coleta apenas os "Não Cadastrado" visíveis com supplier_id conhecido
        to_register = [
            row for r, row in enumerate(self._all_rows)
            if not self._table.isRowHidden(r)
            and row.get("status") == _STATUS_NAO
            and row.get("supplier_id") is not None
        ]
        if not to_register:
            QMessageBox.information(
                self, "Cadastrar",
                "Nenhum produto não cadastrado visível com fornecedor correspondente."
            )
            return

        fornecedores_envolvidos = len({r["supplier_db_name"] for r in to_register})
        resp = QMessageBox.question(
            self, "Cadastrar Produtos",
            f"{len(to_register)} produto(s) serão cadastrados em {fornecedores_envolvidos} fornecedor(es).\n\n"
            "Deseja continuar?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if resp != QMessageBox.Yes:
            return

        self._btn_cadastrar.setEnabled(False)
        self._btn_verificar.setEnabled(False)
        self._btn_export.setEnabled(False)
        self._progress.setValue(0)
        self._progress.show()
        self._progress_label.setText("Iniciando...")
        self._progress_label.show()

        self._c_thread = QThread(self)
        self._c_worker = _CadastrarTodosWorker(to_register, self._operation_type, self._repo)
        self._c_worker.moveToThread(self._c_thread)
        self._c_thread.started.connect(self._c_worker.run)
        self._c_worker.progress.connect(self._on_c_progress)
        self._c_worker.finished.connect(self._on_c_done)
        self._c_worker.failed.connect(self._on_c_failed)
        self._c_worker.finished.connect(self._c_thread.quit)
        self._c_worker.failed.connect(self._c_thread.quit)
        self._c_thread.finished.connect(self._c_worker.deleteLater)
        self._c_thread.finished.connect(self._c_thread.deleteLater)
        self._c_thread.finished.connect(lambda: setattr(self, "_c_thread", None))
        self._c_thread.finished.connect(lambda: setattr(self, "_c_worker", None))
        self._c_thread.start()

    def _on_c_progress(self, current: int, total: int, msg: str) -> None:
        if total > 0:
            self._progress.setValue(int(current / total * 100))
        self._progress_label.setText(msg)

    def _on_c_done(self, stats: dict) -> None:
        self._progress.hide()
        self._progress_label.hide()
        self._btn_verificar.setEnabled(True)
        erros = stats.get("erros", [])
        txt = (
            f"Cadastrados: {stats['cadastrados']} | "
            f"Ignorados: {stats['ignorados']} | "
            f"Erros: {len(erros)}"
        )
        if erros:
            detalhe = "\n".join(f"  [{e['code']}] {e['erro']}" for e in erros[:10])
            txt += f"\n\n{detalhe}"
        QMessageBox.information(self, "Cadastro Concluído", txt)
        # Re-verifica para atualizar os resultados
        self._run_verificar()

    def _on_c_failed(self, msg: str) -> None:
        self._progress.hide()
        self._progress_label.hide()
        self._btn_verificar.setEnabled(True)
        self._btn_cadastrar.setEnabled(True)
        QMessageBox.critical(self, "Erro no Cadastro", f"Erro ao cadastrar:\n{msg}")

    # ── Exportar ──────────────────────────────────────────────────────────────

    def _export(self) -> None:
        if not self._all_rows:
            return

        import re
        op = self._operation_type.lower()
        default_name = f"verificacao_fornecedor_{op}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Salvar Excel", default_name, "Arquivo Excel (*.xlsx)"
        )
        if not path:
            return

        try:
            import os
            from app.exporters.workbook_exporter import write_simple_excel_workbook
            from app.exporters.excel_base import (
                EXCEL_STYLE_STATUS_OK, EXCEL_STYLE_STATUS_NAO, EXCEL_STYLE_STATUS_DIV,
                EXCEL_STYLE_ROW_OK, EXCEL_STYLE_ROW_NAO, EXCEL_STYLE_ROW_DIV,
            )

            # Usa "Sem Correspondência" com o estilo de DIV (divergência)
            _STATUS_STYLE = {
                _STATUS_OK:  (EXCEL_STYLE_STATUS_OK,  EXCEL_STYLE_ROW_OK),
                _STATUS_NAO: (EXCEL_STYLE_STATUS_NAO, EXCEL_STYLE_ROW_NAO),
                _STATUS_SEM: (EXCEL_STYLE_STATUS_DIV, EXCEL_STYLE_ROW_DIV),
            }

            # Exporta apenas os rows visíveis (respeita filtro atual)
            visible_rows = [
                row for r, row in enumerate(self._all_rows)
                if not self._table.isRowHidden(r)
            ]

            all_fields  = _MAIN_FIELDS + _EXPORT_EXTRA_FIELDS
            all_headers = _MAIN_HEADERS + _EXPORT_EXTRA_HEADERS

            data   = [[str(row.get(f) or "") for f in all_fields] for row in visible_rows]
            styles = []
            for row in visible_rows:
                status = str(row.get("status") or "")
                s, r = _STATUS_STYLE.get(status, (EXCEL_STYLE_STATUS_DIV, EXCEL_STYLE_ROW_DIV))
                styles.append([s] + [r] * (len(all_fields) - 1))

            write_simple_excel_workbook(
                Path(path),
                [(
                    f"Verif. Fornecedor {self._operation_type}",
                    all_headers,
                    data,
                    {"row_style_ids": styles, "include_total": False},
                )],
            )
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Exportar")
            msg.setText(f"Exportação concluída.\n\nDeseja abrir o arquivo?\n{path}")
            btn_sim = msg.addButton("Sim", QMessageBox.YesRole)
            msg.addButton("Não", QMessageBox.NoRole)
            msg.exec()
            if msg.clickedButton() == btn_sim:
                os.startfile(path)
        except Exception as exc:
            QMessageBox.critical(self, "Exportar", f"Erro ao exportar:\n{exc}")

    # ── Atualizar Código Fornecedor via XML ────────────────────────────────────

    def _abrir_atualizacao_codigo(self) -> None:
        if self._empresa_id is None:
            QMessageBox.warning(self, "Verificação", "Execute a verificação antes.")
            return
        dlg = _AtualizarCodigoFornecedorDialog(
            rows=self._rows,
            empresa_id=self._empresa_id,
            repository=self._repo,
            parent=self,
        )
        dlg.exec()

    # ── Aba 2: Atualizar Fiscais via XML ───────────────────────────────────────

    def _f_browse_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Selecionar pasta com XMLs das NFe", "")
        if folder:
            self._f_folder_edit.setText(folder)
            self._f_btn_scan.setEnabled(True)

    def _f_run_scan(self) -> None:
        folder = self._f_folder_edit.text().strip()
        if not folder:
            return
        empresa_id = self._empresa_id
        if empresa_id is None:
            empresa_id = self._resolve_empresa_id()
        if not empresa_id:
            QMessageBox.warning(
                self, "Empresa não selecionada",
                "Execute a verificação na aba 1 ou selecione uma empresa primeiro.",
            )
            return

        self._f_btn_scan.setEnabled(False)
        self._f_btn_apply.setEnabled(False)
        self._f_progress.setRange(0, 100)
        self._f_progress.setValue(0)
        self._f_progress.show()
        self._f_status_label.setText("Escaneando XMLs e vinculando produtos...")

        scope_sped_only = self._f_radio_sped.isChecked()

        self._f_thread = QThread(self)
        self._f_worker = _AtualizarFiscaisWorker(
            self._rows, int(empresa_id), self._repo, folder,
            scope_sped_only=scope_sped_only,
        )
        self._f_worker.moveToThread(self._f_thread)
        self._f_thread.started.connect(self._f_worker.run)
        self._f_worker.progress.connect(self._f_on_progress)
        self._f_worker.finished.connect(self._f_on_done)
        self._f_worker.failed.connect(self._f_on_failed)
        self._f_worker.finished.connect(lambda *_: self._f_thread.quit())
        self._f_worker.failed.connect(self._f_thread.quit)
        self._f_thread.finished.connect(self._f_worker.deleteLater)
        self._f_thread.finished.connect(self._f_thread.deleteLater)
        self._f_thread.finished.connect(lambda: setattr(self, "_f_thread", None))
        self._f_thread.finished.connect(lambda: setattr(self, "_f_worker", None))
        self._f_thread.start()

    def _f_on_progress(self, current: int, total: int, msg: str) -> None:
        if total > 0:
            self._f_progress.setValue(int(current / total * 100))
        self._f_status_label.setText(msg)

    def _f_on_done(self, proposals: list[dict], stats: dict) -> None:
        self._f_progress.hide()
        self._f_btn_scan.setEnabled(True)
        self._f_btn_export.setEnabled(bool(proposals))
        self._fiscal_proposals = proposals
        self._f_fill_table()

        n_diff    = stats.get(_FS_DIFF,   0)
        n_ok      = stats.get(_FS_OK,     0)
        n_sem_xml = stats.get(_FS_NO_XML, 0)
        n_sem_cad = stats.get(_FS_NO_CAD, 0)
        n_total   = len(proposals)

        # Atualiza rótulos dos botões de filtro com contagens
        for label, btn in self._f_filter_btns.items():
            if label == "Todos":
                btn.setText(f"Todos ({n_total})")
            elif label == _FS_DIFF:
                btn.setText(f"{_FS_DIFF} ({n_diff})")
            elif label == _FS_OK:
                btn.setText(f"{_FS_OK} ({n_ok})")
            elif label == _FS_NO_XML:
                btn.setText(f"{_FS_NO_XML} ({n_sem_xml})")
            elif label == _FS_NO_CAD:
                btn.setText(f"{_FS_NO_CAD} ({n_sem_cad})")

        parts = []
        if n_diff:
            n_uniq = sum(1 for p in proposals if p.get("status") == _FS_DIFF and p.get("confidence") == "Único")
            n_amb  = n_diff - n_uniq
            parts.append(f"<b>{n_diff} com diferença</b> ({n_uniq} únicos · {n_amb} ambíguos)")
        if n_ok:
            parts.append(f"{n_ok} já corretos")
        if n_sem_xml:
            parts.append(f"{n_sem_xml} sem XML")
        if n_sem_cad:
            parts.append(f"{n_sem_cad} sem cadastro")
        summary = " · ".join(parts) if parts else "Nenhum produto encontrado"

        if n_diff:
            self._f_status_label.setText(summary + ". Selecione campos e clique em Aplicar.")
        else:
            self._f_status_label.setText(
                summary + ". Verifique se os XMLs correspondem às NFes do SPED."
            )
        self._f_update_count()

    def _f_on_failed(self, msg: str) -> None:
        self._f_progress.hide()
        self._f_btn_scan.setEnabled(True)
        QMessageBox.critical(self, "Erro ao Escanear", f"Erro:\n{msg}")

    # ── Tabela fiscais ─────────────────────────────────────────────────────────

    # Cores por status
    _BG_DIFF    = QColor("#c6efce"); _FG_DIFF    = QColor("#276221")
    _BG_SAME    = QColor("#eeeeee"); _FG_SAME    = QColor("#888888")
    _BG_UNICO   = QColor("#c6efce"); _FG_UNICO   = QColor("#276221")
    _BG_AMBIG   = QColor("#ffeb9c"); _FG_AMBIG   = QColor("#7d4a00")
    _BG_SEM_XML = QColor("#fce4b6"); _FG_SEM_XML = QColor("#7d4a00")
    _BG_SEM_CAD = QColor("#e0e0e0"); _FG_SEM_CAD = QColor("#666666")
    _BG_OK      = QColor("#eaf4e8"); _FG_OK      = QColor("#357a38")

    _STATUS_COLORS = {
        _FS_DIFF:  ("#fff9c4", "#7d5200"),
        _FS_OK:    ("#eaf4e8", "#357a38"),
        _FS_NO_XML:("#fce4b6", "#7d4a00"),
        _FS_NO_CAD:("#e0e0e0", "#555555"),
    }

    @staticmethod
    def _fmt_num(v) -> str:
        try: return f"{float(v):.4f}".replace(".", ",")
        except: return "0,0000"

    def _f_field_cell(self, p: dict, field: str) -> tuple[str, bool]:
        status = p.get("status", "")
        if status in (_FS_NO_XML, _FS_NO_CAD):
            raw = p.get(f"cat_{field}")
            if field in _FISCAL_NUMERIC:
                cat_s = self._fmt_num(raw) if raw is not None and raw != "" else ""
            else:
                cat_s = str(raw or "").strip()
            return cat_s, False
        enabled = self._f_field_checks[field].isChecked()
        is_num = field in _FISCAL_NUMERIC
        if is_num:
            cat_v = float(p.get(f"cat_{field}") or 0)
            xml_v = float(p.get(f"xml_{field}") or 0)
            cat_s = self._fmt_num(cat_v)
            xml_s = self._fmt_num(xml_v)
            diff  = abs(cat_v - xml_v) > 1e-4
        else:
            cat_s = str(p.get(f"cat_{field}") or "").strip()
            xml_s = str(p.get(f"xml_{field}") or "").strip()
            diff  = cat_s != xml_s
        if not enabled:
            return cat_s, False
        return (f"{cat_s} → {xml_s}", True) if diff else (cat_s, False)

    def _f_set_cell(self, r: int, c: int, text: str, bg: QColor, fg: QColor | None = None, bold: bool = False) -> None:
        item = self._f_table.item(r, c)
        if item is None:
            item = QTableWidgetItem()
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
            self._f_table.setItem(r, c, item)
        item.setText(str(text)); item.setBackground(bg)
        if fg: item.setForeground(fg)
        f = item.font(); f.setBold(bold); item.setFont(f)

    def _f_set_filter(self, label: str) -> None:
        self._f_filter = label
        for lbl, btn in self._f_filter_btns.items():
            btn.setChecked(lbl == label)
        self._f_fill_table()

    def _f_fill_table(self, *_) -> None:
        if not self._fiscal_proposals:
            self._f_displayed = []
            return
        # Aplica filtro
        f = self._f_filter
        if f == "Todos":
            self._f_displayed = list(self._fiscal_proposals)
        else:
            self._f_displayed = [p for p in self._fiscal_proposals if p.get("status") == f]

        rows = self._f_displayed
        self._f_table.setUpdatesEnabled(False)
        self._f_table.setRowCount(len(rows))

        for r, p in enumerate(rows):
            status = p.get("status", "")
            is_u   = p.get("confidence") == "Único"
            row_bg = _ROW_EVEN if r % 2 == 0 else _ROW_ODD
            can_select = status == _FS_DIFF

            # col 0 — checkbox (só habilitado para "Com diferença")
            chk = self._f_table.item(r, 0)
            if chk is None:
                chk = QTableWidgetItem()
                self._f_table.setItem(r, 0, chk)
            if can_select:
                chk.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable)
                chk.setCheckState(Qt.Checked if p.get("selected") else Qt.Unchecked)
            else:
                chk.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                chk.setText("")
            chk.setBackground(row_bg)

            # col 1 — Status (célula colorida)
            sc = self._STATUS_COLORS.get(status, ("#ffffff", "#000000"))
            self._f_set_cell(r, _STATUS_COL, status,
                             QColor(sc[0]), QColor(sc[1]), bold=True)

            # col 2 — Confiança
            if can_select or status == _FS_OK:
                cbg = self._BG_UNICO if is_u else self._BG_AMBIG
                cfg = self._FG_UNICO if is_u else self._FG_AMBIG
                self._f_set_cell(r, 2, p.get("confidence", ""), cbg, cfg, bold=True)
            else:
                self._f_set_cell(r, 2, "", row_bg)

            # col 3 — Tipo match
            self._f_set_cell(r, 3, p.get("match_type", ""), row_bg)

            # cols 4-7 — Fixas
            self._f_set_cell(r, 4, p.get("fornecedor_nome", ""),   row_bg)
            self._f_set_cell(r, 5, p.get("description", ""),       row_bg)
            self._f_set_cell(r, 6, p.get("codigo_empresa", ""),    row_bg)
            self._f_set_cell(r, 7, p.get("codigo_fornecedor", ""), row_bg)

            # cols fiscais
            for field, col in _FISCAL_FIELD_COLS.items():
                text, diff = self._f_field_cell(p, field)
                if diff:
                    self._f_set_cell(r, col, text, self._BG_DIFF, self._FG_DIFF, bold=True)
                else:
                    self._f_set_cell(r, col, text, self._BG_SAME, self._FG_SAME)

            # col chave NFe
            self._f_set_cell(r, _FISCAL_KEY_COL, p.get("document_key", ""), row_bg)

        self._f_table.resizeColumnsToContents()
        self._f_table.setUpdatesEnabled(True)
        self._f_update_count()

    def _f_refresh_table(self, *_) -> None:
        self._f_fill_table()

    def _f_on_cell_clicked(self, row: int, col: int) -> None:
        if col != 0 or row >= len(self._f_displayed):
            return
        p = self._f_displayed[row]
        if p.get("status") != _FS_DIFF:
            return
        item = self._f_table.item(row, 0)
        if item is None:
            return
        new_state = Qt.Unchecked if item.checkState() == Qt.Checked else Qt.Checked
        item.setCheckState(new_state)
        p["selected"] = (new_state == Qt.Checked)
        self._f_update_count()

    def _f_set_all(self, state: bool) -> None:
        for p in self._fiscal_proposals:
            if p.get("status") == _FS_DIFF:
                p["selected"] = state
        # Atualiza checkboxes visíveis
        for r, p in enumerate(self._f_displayed):
            if p.get("status") == _FS_DIFF:
                item = self._f_table.item(r, 0)
                if item:
                    item.setCheckState(Qt.Checked if state else Qt.Unchecked)
        self._f_update_count()

    def _f_deselect_ambiguous(self) -> None:
        for p in self._fiscal_proposals:
            if p.get("status") == _FS_DIFF and p.get("confidence") != "Único":
                p["selected"] = False
        for r, p in enumerate(self._f_displayed):
            if p.get("status") == _FS_DIFF and p.get("confidence") != "Único":
                item = self._f_table.item(r, 0)
                if item:
                    item.setCheckState(Qt.Unchecked)
        self._f_update_count()

    def _f_update_count(self) -> None:
        n_diff = sum(1 for p in self._fiscal_proposals if p.get("status") == _FS_DIFF)
        sel    = sum(1 for p in self._fiscal_proposals if p.get("selected"))
        enabled = [k for k, chk in self._f_field_checks.items() if chk.isChecked()]
        self._f_count_label.setText(f"{sel}/{n_diff} com diferença selecionados")
        self._f_btn_apply.setEnabled(sel > 0 and bool(enabled))

    # ── Aplicar fiscais ────────────────────────────────────────────────────────

    def _f_apply(self) -> None:
        to_do  = [p for p in self._fiscal_proposals if p.get("selected")]
        fields = [k for k, chk in self._f_field_checks.items() if chk.isChecked()]
        if not to_do or not fields:
            return
        campos_str = ", ".join(
            lbl for k, lbl in (
                ("cst_icms", "CST ICMS"), ("aliquota_icms", "Alíq. ICMS"),
                ("cst_pis",  "CST PIS"),  ("aliquota_pis",  "Alíq. PIS"),
                ("cst_cofins", "CST COFINS"), ("aliquota_cofins", "Alíq. COFINS"),
            ) if k in fields
        )
        resp = QMessageBox.question(
            self, "Aplicar Atualizações Fiscais",
            f"<b>{len(to_do)}</b> produto(s) terão os seguintes campos atualizados "
            f"com os dados do XML:<br><br><b>{campos_str}</b><br><br>"
            "Somente campos onde o valor do XML difere do catálogo serão gravados.<br>"
            "Deseja continuar?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if resp != QMessageBox.Yes:
            return

        self._f_btn_apply.setEnabled(False)
        self._f_btn_scan.setEnabled(False)
        self._f_progress.setRange(0, 100)
        self._f_progress.setValue(0)
        self._f_progress.show()

        self._fa_thread = QThread(self)
        self._fa_worker = _AplicarFiscaisWorker(to_do, self._repo, fields)
        self._fa_worker.moveToThread(self._fa_thread)
        self._fa_thread.started.connect(self._fa_worker.run)
        self._fa_worker.progress.connect(self._f_on_apply_progress)
        self._fa_worker.finished.connect(self._f_on_apply_done)
        self._fa_worker.failed.connect(self._f_on_apply_failed)
        self._fa_worker.finished.connect(self._fa_thread.quit)
        self._fa_worker.failed.connect(self._fa_thread.quit)
        self._fa_thread.finished.connect(self._fa_worker.deleteLater)
        self._fa_thread.finished.connect(self._fa_thread.deleteLater)
        self._fa_thread.finished.connect(lambda: setattr(self, "_fa_thread", None))
        self._fa_thread.finished.connect(lambda: setattr(self, "_fa_worker", None))
        self._fa_thread.start()

    def _f_on_apply_progress(self, current: int, total: int, msg: str) -> None:
        if total > 0:
            self._f_progress.setValue(int(current / total * 100))
        self._f_status_label.setText(msg)

    def _f_on_apply_done(self, stats: dict) -> None:
        self._f_progress.hide()
        self._f_btn_scan.setEnabled(True)
        erros = stats.get("erros", [])
        txt = (
            f"Atualizados: {stats['atualizados']} | "
            f"Sem diferença: {stats.get('sem_diff', 0)} | "
            f"Erros: {len(erros)}"
        )
        if erros:
            txt += "\n" + "\n".join(f"  [id={e['product_id']}] {e['erro']}" for e in erros[:10])
        QMessageBox.information(self, "Atualização Concluída", txt)
        applied = {p["product_id"] for p in self._fiscal_proposals if p.get("selected")}
        self._fiscal_proposals = [p for p in self._fiscal_proposals if p["product_id"] not in applied]
        self._f_fill_table()
        self._f_update_count()
        self._f_status_label.setText(
            f"Atualizados: {stats['atualizados']}. Propostas restantes: {len(self._fiscal_proposals)}"
        )

    def _f_on_apply_failed(self, msg: str) -> None:
        self._f_progress.hide()
        self._f_btn_scan.setEnabled(True)
        self._f_btn_apply.setEnabled(True)
        QMessageBox.critical(self, "Erro ao Aplicar", f"Erro:\n{msg}")

    # ── Exportar fiscais ───────────────────────────────────────────────────────

    def _f_export(self) -> None:
        rows = self._f_displayed if self._f_displayed else self._fiscal_proposals
        if not rows:
            return

        op = self._operation_type.lower()
        default_name = f"atualizar_fiscais_xml_{op}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Salvar Excel", default_name, "Arquivo Excel (*.xlsx)"
        )
        if not path:
            return

        try:
            import os
            from app.exporters.workbook_exporter import write_simple_excel_workbook
            from app.exporters.excel_base import (
                EXCEL_STYLE_STATUS_OK, EXCEL_STYLE_STATUS_NAO, EXCEL_STYLE_STATUS_DIV,
                EXCEL_STYLE_ROW_OK, EXCEL_STYLE_ROW_NAO, EXCEL_STYLE_ROW_DIV,
            )

            _STATUS_STYLE = {
                _FS_DIFF:   (EXCEL_STYLE_STATUS_DIV, EXCEL_STYLE_ROW_DIV),
                _FS_OK:     (EXCEL_STYLE_STATUS_OK,  EXCEL_STYLE_ROW_OK),
                _FS_NO_XML: (EXCEL_STYLE_STATUS_NAO, EXCEL_STYLE_ROW_NAO),
                _FS_NO_CAD: (EXCEL_STYLE_STATUS_NAO, EXCEL_STYLE_ROW_NAO),
            }

            headers = [
                "Status", "Conf.", "Tipo Match", "Fornecedor", "Produto",
                "Cód. Empresa", "Cód. Forn.",
                "CFOP Saída Forn. (Cad)", "CFOP Saída Forn. (XML)",
                "Origem (Cad)", "Origem (XML)",
                "CST ICMS (Cad)", "CST ICMS (XML)",
                "Alíq. ICMS % (Cad)", "Alíq. ICMS % (XML)",
                "Red. BC ICMS % (Cad)", "Red. BC ICMS % (XML)",
                "MVA % (Cad)", "MVA % (XML)",
                "Vl. ICMS-ST (Cad)", "Vl. ICMS-ST (XML)",
                "Alíq. ICMS ST % (Cad)", "Alíq. ICMS ST % (XML)",
                "CST PIS (Cad)", "CST PIS (XML)",
                "Alíq. PIS % (Cad)", "Alíq. PIS % (XML)",
                "CST COFINS (Cad)", "CST COFINS (XML)",
                "Alíq. COFINS % (Cad)", "Alíq. COFINS % (XML)",
                "Chave NFe",
            ]

            def _fmt(v, is_num):
                if is_num:
                    try: return f"{float(v):.4f}".replace(".", ",")
                    except: return ""
                return str(v or "").strip()

            _NUM = _FISCAL_NUMERIC
            data, styles = [], []
            for p in rows:
                status = p.get("status", "")
                row_data = [
                    status,
                    p.get("confidence", ""),
                    p.get("match_type", ""),
                    p.get("fornecedor_nome", ""),
                    p.get("description", ""),
                    p.get("codigo_empresa", ""),
                    p.get("codigo_fornecedor", ""),
                ]
                for f in _FISCAL_FIELD_COLS:
                    is_num = f in _NUM
                    row_data.append(_fmt(p.get(f"cat_{f}"), is_num))
                    row_data.append(_fmt(p.get(f"xml_{f}"), is_num))
                row_data.append(p.get("document_key", ""))
                data.append(row_data)
                s, r = _STATUS_STYLE.get(status, (EXCEL_STYLE_STATUS_DIV, EXCEL_STYLE_ROW_DIV))
                styles.append([s] + [r] * (len(row_data) - 1))

            write_simple_excel_workbook(
                Path(path),
                [(
                    f"Fiscais XML {self._operation_type}",
                    headers,
                    data,
                    {"row_style_ids": styles, "include_total": False},
                )],
            )
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Exportar")
            msg.setText(f"Exportação concluída.\n\nDeseja abrir o arquivo?\n{path}")
            btn_sim = msg.addButton("Sim", QMessageBox.YesRole)
            msg.addButton("Não", QMessageBox.NoRole)
            msg.exec()
            if msg.clickedButton() == btn_sim:
                os.startfile(path)
        except Exception as exc:
            QMessageBox.critical(self, "Exportar", f"Erro ao exportar:\n{exc}")

    # ── Aba 3: Conferência XMLs × SPED ────────────────────────────────────────

    _CONF_BG = {
        "Ausente":    QColor("#f4cccc"),
        "Encontrado": QColor("#c6efce"),
        "Extra":      QColor("#d9e8f5"),
    }
    _CONF_FG = {
        "Ausente":    QColor("#7d0000"),
        "Encontrado": QColor("#276221"),
        "Extra":      QColor("#1a4a7a"),
    }

    def _conf_browse_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Selecionar pasta com XMLs das NFe", "")
        if folder:
            self._conf_folder_edit.setText(folder)
            self._conf_btn_scan.setEnabled(True)

    def _conf_run_scan(self) -> None:
        folder = self._conf_folder_edit.text().strip()
        if not folder:
            return
        if not self._rows:
            QMessageBox.warning(self, "Sem dados SPED",
                "Carregue um arquivo SPED antes de conferir os XMLs.")
            return
        self._conf_btn_scan.setEnabled(False)
        self._conf_btn_export.setEnabled(False)
        self._conf_progress.setRange(0, 100)
        self._conf_progress.setValue(0)
        self._conf_progress.show()
        self._conf_status_label.setText("Escaneando XMLs e cruzando com SPED...")

        self._conf_thread = QThread(self)
        self._conf_worker = _ConferirXMLsWorker(self._rows, folder)
        self._conf_worker.moveToThread(self._conf_thread)
        self._conf_thread.started.connect(self._conf_worker.run)
        self._conf_worker.progress.connect(self._conf_on_progress)
        self._conf_worker.finished.connect(self._conf_on_done)
        self._conf_worker.failed.connect(self._conf_on_failed)
        self._conf_worker.finished.connect(self._conf_thread.quit)
        self._conf_worker.failed.connect(self._conf_thread.quit)
        self._conf_thread.finished.connect(self._conf_worker.deleteLater)
        self._conf_thread.finished.connect(self._conf_thread.deleteLater)
        self._conf_thread.finished.connect(lambda: setattr(self, "_conf_thread", None))
        self._conf_thread.finished.connect(lambda: setattr(self, "_conf_worker", None))
        self._conf_thread.start()

    def _conf_on_progress(self, current: int, total: int, msg: str) -> None:
        if total > 0:
            self._conf_progress.setValue(int(current / total * 100))
        self._conf_status_label.setText(msg)

    def _conf_on_done(self, results: list, stats: dict) -> None:
        self._conf_progress.hide()
        self._conf_btn_scan.setEnabled(True)
        self._conf_all_results = results
        # Atualiza métricas
        for key, val in stats.items():
            if key in self._conf_metric_labels:
                self._conf_metric_labels[key].setText(str(val))
        # Atualiza filtros com contagens
        self._conf_filter_btns["Todos"].setText(f"Todos ({len(results)})")
        self._conf_filter_btns["Ausente"].setText(f"Ausente ({stats['ausente']})")
        self._conf_filter_btns["Encontrado"].setText(f"Encontrado ({stats['encontrado']})")
        self._conf_filter_btns["Extra"].setText(f"Extra ({stats['extra']})")
        self._conf_fill_table()
        self._conf_btn_export.setEnabled(bool(results))
        self._conf_status_label.setText(
            f"Total: {len(results)} | "
            f"Encontrados: {stats['encontrado']} | "
            f"Ausentes no disco: {stats['ausente']} | "
            f"Extras: {stats['extra']}"
        )

    def _conf_on_failed(self, msg: str) -> None:
        self._conf_progress.hide()
        self._conf_btn_scan.setEnabled(True)
        QMessageBox.critical(self, "Erro na Conferência", f"Erro:\n{msg}")

    def _conf_set_filter(self, mode: str) -> None:
        self._conf_filter = mode
        for m, btn in self._conf_filter_btns.items():
            btn.setChecked(m == mode)
        self._conf_fill_table()

    def _conf_fill_table(self) -> None:
        rows = self._conf_all_results
        if self._conf_filter != "Todos":
            rows = [r for r in rows if r["status"] == self._conf_filter]

        self._conf_table.setUpdatesEnabled(False)
        self._conf_table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            status = r["status"]
            bg = self._CONF_BG.get(status, _ROW_EVEN)
            fg = self._CONF_FG.get(status, QColor("#000000"))
            row_bg = bg if i % 2 == 0 else bg.lighter(103)
            cells = [
                status,
                r.get("number", ""),
                r.get("date", ""),
                r.get("issuer_cnpj", ""),
                r.get("issuer_name", ""),
                f"R$ {r.get('total_value', 0):.2f}".replace(".", ",") if r.get("total_value") else "",
                r.get("participant", ""),
                r.get("document_key", ""),
                r.get("xml_file", ""),
            ]
            for col, text in enumerate(cells):
                item = self._conf_table.item(i, col)
                if item is None:
                    item = QTableWidgetItem()
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                    self._conf_table.setItem(i, col, item)
                item.setText(str(text))
                item.setBackground(row_bg)
                item.setForeground(fg)
                f = item.font(); f.setBold(col == 0); item.setFont(f)
        self._conf_table.resizeColumnsToContents()
        self._conf_table.setUpdatesEnabled(True)

    def _conf_export(self) -> None:
        if not self._conf_all_results:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Exportar Conferência XML × SPED", "conferencia_xml_sped.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Conferência XML × SPED"

            headers = ["Status", "Número NF", "Data", "CNPJ Emitente", "Emitente",
                       "Valor Total", "Participante SPED", "Chave NFe", "Arquivo XML"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)

            _fill = {
                "Ausente":    PatternFill("solid", fgColor="F4CCCC"),
                "Encontrado": PatternFill("solid", fgColor="C6EFCE"),
                "Extra":      PatternFill("solid", fgColor="D9E8F5"),
            }
            for r in self._conf_all_results:
                row_data = [
                    r.get("status", ""),
                    r.get("number", ""),
                    r.get("date", ""),
                    r.get("issuer_cnpj", ""),
                    r.get("issuer_name", ""),
                    r.get("total_value", 0),
                    r.get("participant", ""),
                    r.get("document_key", ""),
                    r.get("xml_file", ""),
                ]
                ws.append(row_data)
                fill = _fill.get(r.get("status", ""))
                if fill:
                    for cell in ws[ws.max_row]:
                        cell.fill = fill

            # Auto-width nas primeiras colunas
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            wb.save(path)
            msg = QMessageBox(self)
            msg.setWindowTitle("Exportar")
            msg.setText(f"Exportação concluída.\n\n{path}")
            btn_sim = msg.addButton("Abrir", QMessageBox.YesRole)
            msg.addButton("Fechar", QMessageBox.NoRole)
            msg.exec()
            if msg.clickedButton() == btn_sim:
                import os; os.startfile(path)
        except Exception as exc:
            QMessageBox.critical(self, "Erro ao Exportar", f"Erro:\n{exc}")


# ── Worker: vincula todos os produtos do catálogo ao XML e compara fiscais ────

class _AtualizarFiscaisWorker(QObject):
    """
    Para cada produto do catálogo da empresa tenta encontrar o item correspondente
    nos XMLs das NFe usando duas estratégias:
      1. Por codigo_fornecedor = cProd do XML (mais confiável)
      2. Por quantidade + valor aproximados dentro da mesma NFe
    Retorna propostas de atualização dos campos fiscais do catálogo.
    """
    progress = Signal(int, int, str)
    finished = Signal(list, dict)   # (proposals, stats)
    failed   = Signal(str)

    _TOL_QTY = Decimal("0.001")
    _TOL_VAL = Decimal("0.01")
    _PRIORITY = {"Código + Qtd/Val": 3, "Por Código": 2, "Por Qtd/Val": 1, "Divergente": 0}

    def __init__(
        self,
        rows: list[dict],
        empresa_id: int,
        repository: MysqlCadastroRepository,
        xml_folder: str,
        scope_sped_only: bool = True,
    ) -> None:
        super().__init__()
        self._rows            = rows
        self._empresa_id      = empresa_id
        self._repo            = repository
        self._xml_folder      = xml_folder
        self._scope_sped_only = scope_sped_only

    def run(self) -> None:
        try:
            from app.parsers.compare_xml import parse_compare_xml_file
            from collections import defaultdict

            def _f(v): return float(v or 0)
            def _s(v): return str(v or "").strip()
            def _digits(s): return "".join(c for c in str(s or "") if c.isdigit())
            TOL = 1e-4

            # 1. Carrega catálogo e monta índices
            self.progress.emit(0, 100, "Carregando catálogo...")
            catalog = self._repo.get_catalog_products_by_empresa_id(self._empresa_id)

            cat_by_emp_cnpj: dict[tuple, list[dict]] = defaultdict(list)
            cat_by_emp:      dict[str,   list[dict]] = defaultdict(list)
            cat_by_forn:     dict[str,   list[dict]] = defaultdict(list)
            cat_by_ean:      dict[str,   list[dict]] = defaultdict(list)
            for p in catalog:
                cfe  = _s(p.get("codigo_empresa"))
                cff  = _s(p.get("codigo_fornecedor"))
                cnpj = _digits(p.get("fornecedor_cnpj"))
                ean  = _digits(p.get("ean"))
                if cfe:
                    cat_by_emp[cfe].append(p)
                    if cnpj:
                        cat_by_emp_cnpj[(cfe, cnpj)].append(p)
                if cff and cff != cfe:
                    cat_by_forn[cff].append(p)
                if ean:
                    cat_by_ean[ean].append(p)

            # 2. Coleta todos os itens únicos do SPED: (code, ptax) → {pname, doc_keys}
            sped_entries: dict[tuple, dict] = {}
            for row in self._rows:
                for detail in row.get("launch_details", []):
                    code    = _s(detail.get("code"))
                    doc_key = _s(detail.get("document_key"))
                    ptax    = _digits(detail.get("participant_tax_id"))
                    pname   = _s(detail.get("participant_name"))
                    if not code or len(doc_key) != 44:
                        continue
                    k = (code, ptax)
                    if k not in sped_entries:
                        sped_entries[k] = {"code": code, "ptax": ptax, "pname": pname, "doc_keys": {}}
                    if doc_key not in sped_entries[k]["doc_keys"]:
                        sped_entries[k]["doc_keys"][doc_key] = detail

            # 3. Escaneia TODOS os XMLs
            xml_files   = sorted(Path(self._xml_folder).rglob("*.xml"))
            total_files = len(xml_files)
            xml_by_key:         dict[str, object] = {}
            xml_items_by_cprod: dict[str, list]   = defaultdict(list)

            for i, xf in enumerate(xml_files, 1):
                self.progress.emit(i, total_files, f"Lendo XML {i}/{total_files}: {xf.name[:50]}")
                try:
                    invoice = parse_compare_xml_file(xf)
                    if invoice:
                        xml_by_key[invoice.key] = invoice
                        for xi in invoice.items:
                            cprod = _s(xi.code)
                            if cprod:
                                xml_items_by_cprod[cprod].append((invoice.key, xi))
                except Exception:
                    pass

            # ── Helpers ─────────────────────────────────────────────────────────

            _NUM_FIELDS = _FISCAL_NUMERIC

            def _cat_fields(cat_prod):
                """Retorna dict cat_{field} + xml_{field} vazio (para sem_xml e sem_cad)."""
                r = {}
                for f in _FISCAL_FIELD_COLS:
                    r[f"cat_{f}"] = _f(cat_prod.get(f)) if f in _NUM_FIELDS else _s(cat_prod.get(f))
                    r[f"xml_{f}"] = 0.0 if f in _NUM_FIELDS else ""
                return r

            def _build(cat_prod, xml_item, mtype, conf, key) -> dict:
                """Monta proposta comparando catálogo × XML; determina status."""
                pid = int(cat_prod["id"])
                c_cfop  = _s(cat_prod.get("cfop_saida_fornecedor")); x_cfop  = _s(xml_item.cfop)
                c_orig  = _s(cat_prod.get("origem_entrada"));         x_orig  = _s(xml_item.orig_icms)
                c_cst   = _s(cat_prod.get("cst_icms"));               x_cst   = _s(xml_item.cst_icms)
                c_aliq  = _f(cat_prod.get("aliquota_icms"));          x_aliq  = _f(xml_item.aliq_icms)     / 100
                c_red   = _f(cat_prod.get("reducao_bc_icms"));        x_red   = _f(xml_item.reducao_bc_icms) / 100
                c_mva   = _f(cat_prod.get("mva"));                    x_mva   = _f(xml_item.mva)
                c_vst   = _f(cat_prod.get("valor_icms_st"));          x_vst   = _f(xml_item.vl_icms_st)
                c_ast   = _f(cat_prod.get("aliquota_icms_st"));       x_ast   = _f(xml_item.aliq_icms_st)  / 100
                c_cp    = _s(cat_prod.get("cst_pis"));                x_cp    = _s(xml_item.cst_pis)
                c_ap    = _f(cat_prod.get("aliquota_pis"));           x_ap    = _f(xml_item.aliq_pis)      / 100
                c_cc    = _s(cat_prod.get("cst_cofins"));             x_cc    = _s(xml_item.cst_cofins)
                c_ac    = _f(cat_prod.get("aliquota_cofins"));        x_ac    = _f(xml_item.aliq_cofins)   / 100
                has_diff = (
                    (x_cfop and c_cfop != x_cfop) or (x_orig and c_orig != x_orig)
                    or c_cst != x_cst or abs(c_aliq - x_aliq) > TOL
                    or abs(c_red - x_red) > TOL or abs(c_mva - x_mva) > TOL
                    or abs(c_vst - x_vst) > 0.005 or abs(c_ast - x_ast) > TOL
                    or c_cp != x_cp or abs(c_ap - x_ap) > TOL
                    or c_cc != x_cc or abs(c_ac - x_ac) > TOL
                )
                status = _FS_DIFF if has_diff else _FS_OK
                return {
                    "status":            status,
                    "selected":          has_diff and conf == "Único",
                    "confidence":        conf,
                    "match_type":        mtype,
                    "product_id":        pid,
                    "fornecedor_nome":   _s(cat_prod.get("fornecedor_nome")),
                    "description":       _s(cat_prod.get("descricao")),
                    "codigo_empresa":    _s(cat_prod.get("codigo_empresa")),
                    "codigo_fornecedor": _s(cat_prod.get("codigo_fornecedor")),
                    "document_key":      key,
                    "cat_cfop_saida_fornecedor": c_cfop, "xml_cfop_saida_fornecedor": x_cfop,
                    "cat_origem_entrada":  c_orig,        "xml_origem_entrada":   x_orig,
                    "cat_cst_icms":        c_cst,         "xml_cst_icms":         x_cst,
                    "cat_aliquota_icms":   c_aliq,        "xml_aliquota_icms":    x_aliq,
                    "cat_reducao_bc_icms": c_red,         "xml_reducao_bc_icms":  x_red,
                    "cat_mva":             c_mva,         "xml_mva":              x_mva,
                    "cat_valor_icms_st":   c_vst,         "xml_valor_icms_st":    x_vst,
                    "cat_aliquota_icms_st":c_ast,         "xml_aliquota_icms_st": x_ast,
                    "cat_cst_pis":         c_cp,          "xml_cst_pis":          x_cp,
                    "cat_aliquota_pis":    c_ap,          "xml_aliquota_pis":     x_ap,
                    "cat_cst_cofins":      c_cc,          "xml_cst_cofins":       x_cc,
                    "cat_aliquota_cofins": c_ac,          "xml_aliquota_cofins":  x_ac,
                }

            def _best_xml(cat_prod, doc_keys: dict):
                """Retorna (xml_item, mtype, conf, key) ou None se não encontrou."""
                bx, bm, bc, bk = None, "", "", ""
                for doc_key, detail in doc_keys.items():
                    invoice = xml_by_key.get(doc_key)
                    if invoice is None:
                        continue
                    cff = _s(cat_prod.get("codigo_fornecedor"))
                    cfe = _s(cat_prod.get("codigo_empresa"))
                    sq  = Decimal(str(detail.get("quantity") or 0))
                    sv  = Decimal(str(detail.get("sale_value") or 0))
                    cff_s = cff.lstrip("0")
                    bc_ = ([xi for xi in invoice.items
                            if _s(xi.code) == cff
                            or (cff_s and _s(xi.code).lstrip("0") == cff_s)]
                           if cff and cff != cfe else [])
                    bv_ = [xi for xi in invoice.items
                           if abs(Decimal(str(xi.quantity)) - sq) <= self._TOL_QTY
                           and abs(Decimal(str(xi.value)) - sv) <= self._TOL_VAL]
                    if bc_ and bv_ and {_s(xi.code) for xi in bc_} & {_s(xi.code) for xi in bv_}:
                        m, c, xi = "Código + Qtd/Val", "Único",  bc_[0]
                    elif bc_ and bv_:
                        m, c, xi = "Divergente",       "Ambíguo", bc_[0]
                    elif bc_:
                        m, c, xi = "Por Código",  ("Único" if len(bc_) == 1 else "Ambíguo"), bc_[0]
                    elif bv_:
                        m, c, xi = "Por Qtd/Val", ("Único" if len(bv_) == 1 else "Ambíguo"), bv_[0]
                    else:
                        continue
                    if self._PRIORITY.get(m, -1) > self._PRIORITY.get(bm, -1):
                        bx, bm, bc, bk = xi, m, c, doc_key
                return (bx, bm, bc, bk) if bx else None

            # 4. Processa cada entrada SPED única
            proposals: list[dict] = []
            seen_pids: set[int]   = set()
            stats = {_FS_DIFF: 0, _FS_OK: 0, _FS_NO_XML: 0, _FS_NO_CAD: 0}
            n_entries = len(sped_entries)

            for idx, ((code, ptax), entry) in enumerate(sped_entries.items()):
                if idx % 100 == 0:
                    self.progress.emit(
                        total_files + idx, total_files + n_entries,
                        f"Analisando {idx + 1}/{n_entries}..."
                    )

                # Encontra catálogo (4 estratégias, igual ao confronto)
                code_digits = _digits(code)
                cats = (cat_by_emp_cnpj.get((code, ptax))
                        or cat_by_emp.get(code)
                        or cat_by_forn.get(code, [])
                        or (cat_by_ean.get(code_digits) if code_digits else []))

                if not cats:
                    doc_key = next(iter(entry["doc_keys"]))
                    proposals.append({
                        "status": _FS_NO_CAD, "selected": False,
                        "confidence": "", "match_type": "Sem cadastro",
                        "product_id": None,
                        "fornecedor_nome":   entry["pname"],
                        "description":       "",
                        "codigo_empresa":    code,
                        "codigo_fornecedor": "",
                        "document_key":      doc_key,
                        **{f"cat_{f}": "" for f in _FISCAL_FIELD_COLS},
                        **{f"xml_{f}": "" for f in _FISCAL_FIELD_COLS},
                    })
                    stats[_FS_NO_CAD] += 1
                    continue

                cat_prod = cats[0]
                pid = int(cat_prod["id"])
                if pid in seen_pids:
                    continue
                seen_pids.add(pid)

                result = _best_xml(cat_prod, entry["doc_keys"])
                if result is None:
                    doc_key = next(iter(entry["doc_keys"]))
                    proposals.append({
                        "status": _FS_NO_XML, "selected": False,
                        "confidence": "", "match_type": "Sem XML",
                        "product_id": pid,
                        "fornecedor_nome":   _s(cat_prod.get("fornecedor_nome")),
                        "description":       _s(cat_prod.get("descricao")),
                        "codigo_empresa":    _s(cat_prod.get("codigo_empresa")),
                        "codigo_fornecedor": _s(cat_prod.get("codigo_fornecedor")),
                        "document_key":      doc_key,
                        **_cat_fields(cat_prod),
                    })
                    stats[_FS_NO_XML] += 1
                    continue

                xml_item, mtype, conf, key = result
                prop = _build(cat_prod, xml_item, mtype, conf, key)
                proposals.append(prop)
                stats[prop["status"]] += 1

            # Fase B — catálogo inteiro (escopo "todo o cadastro")
            if not self._scope_sped_only:
                self.progress.emit(total_files, total_files,
                                   "Fase B: catálogo sem vínculo SPED...")
                for cat_prod in catalog:
                    pid = int(cat_prod["id"])
                    if pid in seen_pids:
                        continue
                    cff = _s(cat_prod.get("codigo_fornecedor"))
                    cfe = _s(cat_prod.get("codigo_empresa"))
                    if not cff or cff == cfe:
                        continue
                    matches = xml_items_by_cprod.get(cff, [])
                    if not matches:
                        continue
                    ref = matches[0][1]
                    all_same = all(
                        _s(xi.cst_icms) == _s(ref.cst_icms)
                        and abs(_f(xi.aliq_icms) - _f(ref.aliq_icms)) < 0.01
                        for _, xi in matches[1:]
                    )
                    conf = "Único" if len(matches) == 1 or all_same else "Ambíguo"
                    key, xi = matches[0]
                    seen_pids.add(pid)
                    prop = _build(cat_prod, xi, "Por Cód. Forn.", conf, key)
                    proposals.append(prop)
                    stats[prop["status"]] += 1

            self.finished.emit(proposals, stats)
        except Exception as exc:
            import traceback
            self.failed.emit(str(exc) + "\n" + traceback.format_exc())


# ── Worker: aplica atualizações fiscais aprovadas ─────────────────────────────

class _AplicarFiscaisWorker(QObject):
    progress = Signal(int, int, str)
    finished = Signal(dict)
    failed   = Signal(str)

    # Campos numéricos — comparados como float; DB e XML usam mesma unidade (%, R$)
    _NUMERIC = _FISCAL_NUMERIC

    def __init__(
        self,
        proposals: list[dict],
        repository: MysqlCadastroRepository,
        enabled_fields: list[str],
    ) -> None:
        super().__init__()
        self._proposals      = proposals
        self._repo           = repository
        self._enabled_fields = enabled_fields

    def run(self) -> None:
        try:
            stats = {"atualizados": 0, "sem_diff": 0, "erros": []}
            to_do = [p for p in self._proposals if p.get("selected")]
            total = len(to_do)
            for idx, p in enumerate(to_do):
                self.progress.emit(idx, total, f"Atualizando {idx + 1}/{total}...")
                try:
                    fields: dict = {}
                    for field in self._enabled_fields:
                        cat_v = p.get(f"cat_{field}")
                        xml_v = p.get(f"xml_{field}")
                        if xml_v is None:
                            continue
                        if field in self._NUMERIC:
                            db_new = float(xml_v)  # xml_ já convertido para decimal (÷100) em _build
                            if abs(float(cat_v or 0) - db_new) > 1e-4:
                                fields[field] = db_new
                        else:
                            cs, xs = str(cat_v or "").strip(), str(xml_v or "").strip()
                            if xs and xs != cs:
                                fields[field] = xs
                    if fields:
                        self._repo.update_produto_campos_xml(p["product_id"], fields)
                        stats["atualizados"] += 1
                    else:
                        stats["sem_diff"] += 1
                except Exception as exc:
                    stats["erros"].append({"product_id": p["product_id"], "erro": str(exc)})
            self.progress.emit(total, total, "Concluído.")
            self.finished.emit(stats)
        except Exception as exc:
            self.failed.emit(str(exc))


# ── Worker: conferência XMLs × SPED ──────────────────────────────────────────

class _ConferirXMLsWorker(QObject):
    """
    Coleta todas as chaves de NFe do SPED (via launch_details), escaneia a pasta
    de XMLs e cruza os dois conjuntos:
      - Ausente   : chave no SPED mas XML não encontrado no disco
      - Encontrado: chave no SPED e XML localizado
      - Extra     : XML no disco mas chave não está no SPED carregado
    """
    progress = Signal(int, int, str)
    finished = Signal(list, dict)   # (results, stats)
    failed   = Signal(str)

    def __init__(self, rows: list[dict], xml_folder: str) -> None:
        super().__init__()
        self._rows       = rows
        self._xml_folder = xml_folder

    def run(self) -> None:
        try:
            from app.parsers.compare_xml import parse_compare_xml_file

            # 1. Coleta chaves SPED (únicas) com info do participante
            sped_key_info: dict[str, str] = {}   # key → participant_name
            for row in self._rows:
                pname = str(row.get("participant_name") or "").strip()
                for detail in row.get("launch_details", []):
                    key = str(detail.get("document_key") or "").strip()
                    if len(key) == 44 and key not in sped_key_info:
                        sped_key_info[key] = pname

            # 2. Escaneia XMLs da pasta
            xml_files = sorted(Path(self._xml_folder).rglob("*.xml"))
            total = len(xml_files)
            xml_by_key: dict[str, object] = {}
            for i, xf in enumerate(xml_files, 1):
                self.progress.emit(i, total, f"Lendo {i}/{total}: {xf.name[:50]}")
                try:
                    invoice = parse_compare_xml_file(xf)
                    if invoice and len(str(invoice.key or "")) == 44:
                        k = str(invoice.key).strip()
                        if k not in xml_by_key:
                            xml_by_key[k] = invoice
                except Exception:
                    pass

            # 3. Cruza os dois conjuntos
            results: list[dict] = []

            for key, pname in sped_key_info.items():
                if key in xml_by_key:
                    inv = xml_by_key[key]
                    results.append({
                        "status":        "Encontrado",
                        "document_key":  key,
                        "number":        str(getattr(inv, "number", "") or "").strip(),
                        "date":          str(getattr(inv, "issue_date", "") or "").strip(),
                        "issuer_cnpj":   str(getattr(inv, "issuer_cnpj", "") or "").strip(),
                        "issuer_name":   str(getattr(inv, "issuer_name", "") or "").strip(),
                        "total_value":   float(getattr(inv, "total_doc", 0) or 0),
                        "participant":   pname,
                        "xml_file":      str(getattr(inv, "file_path", "") or "").strip(),
                    })
                else:
                    results.append({
                        "status":        "Ausente",
                        "document_key":  key,
                        "number":        "",
                        "date":          "",
                        "issuer_cnpj":   "",
                        "issuer_name":   "",
                        "total_value":   0.0,
                        "participant":   pname,
                        "xml_file":      "",
                    })

            for key, inv in xml_by_key.items():
                if key not in sped_key_info:
                    results.append({
                        "status":        "Extra",
                        "document_key":  key,
                        "number":        str(getattr(inv, "number", "") or "").strip(),
                        "date":          str(getattr(inv, "issue_date", "") or "").strip(),
                        "issuer_cnpj":   str(getattr(inv, "issuer_cnpj", "") or "").strip(),
                        "issuer_name":   str(getattr(inv, "issuer_name", "") or "").strip(),
                        "total_value":   float(getattr(inv, "total_doc", 0) or 0),
                        "participant":   "",
                        "xml_file":      str(getattr(inv, "file_path", "") or "").strip(),
                    })

            enc  = sum(1 for r in results if r["status"] == "Encontrado")
            aus  = sum(1 for r in results if r["status"] == "Ausente")
            ext  = sum(1 for r in results if r["status"] == "Extra")
            stats = {
                "sped": len(sped_key_info),
                "xmls": len(xml_by_key),
                "encontrado": enc,
                "ausente":    aus,
                "extra":      ext,
            }
            # Ordena: Ausentes primeiro, depois Encontrados, depois Extras
            _ord = {"Ausente": 0, "Encontrado": 1, "Extra": 2}
            results.sort(key=lambda r: (_ord.get(r["status"], 9), r.get("date", ""), r.get("number", "")))
            self.finished.emit(results, stats)
        except Exception as exc:
            import traceback
            self.failed.emit(str(exc) + "\n" + traceback.format_exc())


# ── Worker: escanear XMLs e propor atualizações de codigo_fornecedor ──────────

class _ScanarXMLsWorker(QObject):
    progress = Signal(int, int, str)
    finished = Signal(list)
    failed   = Signal(str)

    _TOL_QTY = Decimal("0.001")
    _TOL_VAL = Decimal("0.01")

    def __init__(
        self,
        rows: list[dict],
        empresa_id: int,
        repository: MysqlCadastroRepository,
        xml_folder: str,
    ) -> None:
        super().__init__()
        self._rows       = rows
        self._empresa_id = empresa_id
        self._repo       = repository
        self._xml_folder = xml_folder

    def run(self) -> None:
        try:
            from app.parsers.compare_xml import parse_compare_xml_file

            # 1. Carrega catálogo da empresa; só interessa produtos auto-preenchidos
            #    (codigo_fornecedor == codigo_empresa → foram cadastrados pelo sistema)
            self.progress.emit(0, 100, "Carregando catálogo da empresa...")
            catalog = self._repo.get_catalog_products_by_empresa_id(self._empresa_id)
            auto_by_code: dict[str, dict] = {}
            for p in catalog:
                cfe = str(p.get("codigo_empresa") or "").strip()
                cff = str(p.get("codigo_fornecedor") or "").strip()
                if not cfe or cfe != cff:
                    continue
                if cfe not in auto_by_code:
                    auto_by_code[cfe] = p
                ean = "".join(c for c in str(p.get("ean") or "") if c.isdigit())
                if ean and ean not in auto_by_code:
                    auto_by_code[ean] = p

            if not auto_by_code:
                self.finished.emit([])
                return

            # 2. Coleta itens SPED para produtos auto-preenchidos, agrupados por chave NFe
            sped_items_by_key: dict[str, list[dict]] = {}
            for row in self._rows:
                for detail in row.get("launch_details", []):
                    code    = str(detail.get("code") or "").strip()
                    doc_key = str(detail.get("document_key") or "").strip()
                    if not code or len(doc_key) != 44:
                        continue
                    cat_prod = auto_by_code.get(code)
                    if cat_prod is None:
                        digits = "".join(c for c in code if c.isdigit())
                        cat_prod = auto_by_code.get(digits) if digits else None
                    if cat_prod is None:
                        continue
                    qty = detail.get("quantity")
                    val = detail.get("sale_value")
                    if qty is None or val is None:
                        continue
                    sped_items_by_key.setdefault(doc_key, []).append({
                        "code":             code,
                        "product_id":       int(cat_prod["id"]),
                        "fornecedor_id":    int(cat_prod["fornecedor_id"]),
                        "code_old":         str(cat_prod.get("codigo_fornecedor") or code),
                        "description_sped": str(detail.get("description") or "").strip(),
                        "ncm_sped":         str(detail.get("ncm") or cat_prod.get("ncm") or ""),
                        "participant_name":  str(detail.get("participant_name") or "").strip(),
                        "fornecedor_nome":  str(cat_prod.get("fornecedor_nome") or "").strip(),
                        "quantity":         Decimal(str(qty)),
                        "sale_value":       Decimal(str(val)),
                        "document_key":     doc_key,
                    })

            if not sped_items_by_key:
                self.finished.emit([])
                return

            # 3. Escaneia XMLs da pasta, indexa por chave de acesso
            needed_keys = set(sped_items_by_key.keys())
            xml_files   = sorted(Path(self._xml_folder).rglob("*.xml"))
            total_files = len(xml_files)
            xml_by_key: dict[str, object] = {}

            for i, xml_file in enumerate(xml_files):
                if not needed_keys:
                    break
                self.progress.emit(i, total_files, f"Lendo: {xml_file.name[:50]}")
                try:
                    invoice = parse_compare_xml_file(xml_file)
                    if invoice and invoice.key in needed_keys:
                        xml_by_key[invoice.key] = invoice
                        needed_keys.discard(invoice.key)
                except Exception:
                    pass

            # 4. Matching SPED ↔ XML por quantidade + valor dentro da mesma NFe
            proposals: list[dict] = []
            seen: set[tuple] = set()  # evita duplicatas (product_id, code_new)

            for doc_key, sped_items in sped_items_by_key.items():
                invoice = xml_by_key.get(doc_key)
                if invoice is None:
                    continue

                for sitem in sped_items:
                    sped_qty = sitem["quantity"]
                    sped_val = sitem["sale_value"]

                    matches = [
                        xi for xi in invoice.items
                        if abs(Decimal(str(xi.quantity)) - sped_qty) <= self._TOL_QTY
                        and abs(Decimal(str(xi.value)) - sped_val) <= self._TOL_VAL
                    ]
                    if not matches:
                        continue

                    # Pega o primeiro match XML (se houver mais de um item com mesma
                    # qty+valor na NF é inconclusivo no nível do documento, mas a
                    # ambiguidade real é detectada depois em nível de fornecedor)
                    xml_item = matches[0]
                    new_code = str(xml_item.code or "").strip()
                    if not new_code or new_code == sitem["code_old"]:
                        continue

                    key_uniq = (sitem["product_id"], new_code)
                    if key_uniq in seen:
                        continue
                    seen.add(key_uniq)

                    proposals.append({
                        "selected":         True,   # será revisado no pós-processamento
                        "confidence":       "Único",
                        "product_id":       sitem["product_id"],
                        "fornecedor_id":    sitem["fornecedor_id"],
                        "fornecedor_nome":  sitem["fornecedor_nome"],
                        "code_old":         sitem["code_old"],
                        "code_new":         new_code,
                        "description_sped": sitem["description_sped"],
                        "description_xml":  str(xml_item.description or "").strip(),
                        "ean_xml":          str(xml_item.ean or ""),
                        "ncm_sped":         sitem["ncm_sped"],
                        "ncm_xml":          str(xml_item.ncm or ""),
                        "quantity":         str(sped_qty),
                        "value":            str(sped_val),
                        "document_key":     doc_key,
                    })

            # Pós-processamento: ambiguidade real = mesmo code_new para produtos
            # diferentes do MESMO fornecedor (cada fornecedor tem seu espaço de códigos,
            # mas se ele repetiu o código para produtos distintos não dá para saber qual é qual)
            from collections import defaultdict
            forn_code_products: dict[tuple[int, str], set[int]] = defaultdict(set)
            for p in proposals:
                forn_code_products[(p["fornecedor_id"], p["code_new"])].add(p["product_id"])

            for p in proposals:
                if len(forn_code_products[(p["fornecedor_id"], p["code_new"])]) > 1:
                    p["confidence"] = "Ambíguo"
                    p["selected"]   = False

            self.finished.emit(proposals)
        except Exception as exc:
            import traceback
            self.failed.emit(str(exc) + "\n" + traceback.format_exc())


# ── Worker: aplicar as atualizações aprovadas ─────────────────────────────────

class _AplicarCodigosWorker(QObject):
    progress = Signal(int, int, str)
    finished = Signal(dict)
    failed   = Signal(str)

    def __init__(
        self,
        proposals: list[dict],
        repository: MysqlCadastroRepository,
    ) -> None:
        super().__init__()
        self._proposals = proposals
        self._repo      = repository

    def run(self) -> None:
        try:
            stats  = {"atualizados": 0, "pulados": 0, "erros": []}
            to_do  = [p for p in self._proposals if p.get("selected")]
            total  = len(to_do)
            for idx, proposal in enumerate(to_do):
                self.progress.emit(idx, total, f"Atualizando {idx + 1}/{total}...")
                forn_id  = proposal.get("fornecedor_id")
                code_new = proposal["code_new"]
                try:
                    # Verifica se o código novo já existe para o mesmo fornecedor
                    if forn_id:
                        existing = self._repo.fetch_supplier_product_by_fornecedor_code(
                            forn_id, code_new
                        )
                        if existing and int(existing["id"]) != int(proposal["product_id"]):
                            stats["pulados"] += 1
                            stats["erros"].append({
                                "code": proposal["code_old"],
                                "erro": (
                                    f"Código '{code_new}' já existe para este fornecedor "
                                    f"(produto id={existing['id']}, "
                                    f"cód.empresa={existing.get('codigo_empresa', '')})"
                                ),
                            })
                            continue
                    self._repo.update_produto_codigo_fornecedor(
                        proposal["product_id"],
                        code_new,
                    )
                    stats["atualizados"] += 1
                except Exception as exc:
                    stats["erros"].append({"code": proposal["code_old"], "erro": str(exc)})
            self.progress.emit(total, total, "Concluído.")
            self.finished.emit(stats)
        except Exception as exc:
            self.failed.emit(str(exc))


# ── Diálogo: preview e aplicação de atualizações de codigo_fornecedor ─────────

class _AtualizarCodigoFornecedorDialog(QDialog):
    """
    Escaneia uma pasta de XMLs das NFe, faz matching com os itens do SPED
    por quantidade + valor, e propõe atualizações de codigo_fornecedor no
    catálogo. Exibe preview com seleção antes de gravar qualquer coisa.
    """

    _PROPOSAL_HEADERS = [
        "✓", "Confiança", "Fornecedor", "Cód. Atual", "Cód. Novo (XML)",
        "Descrição SPED", "Descrição XML", "EAN XML", "NCM SPED", "NCM XML",
        "Qtd.", "Valor", "Chave NFe",
    ]
    _PROPOSAL_FIELDS = [
        None, "confidence", "fornecedor_nome", "code_old", "code_new",
        "description_sped", "description_xml", "ean_xml", "ncm_sped", "ncm_xml",
        "quantity", "value", "document_key",
    ]

    _BG_UNICO  = QColor("#c6efce")
    _FG_UNICO  = QColor("#276221")
    _BG_AMBIG  = QColor("#ffeb9c")
    _FG_AMBIG  = QColor("#7d4a00")

    def __init__(
        self,
        rows: list[dict],
        empresa_id: int,
        repository: MysqlCadastroRepository,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._rows       = rows
        self._empresa_id = empresa_id
        self._repo       = repository
        self._proposals: list[dict] = []
        self._s_thread: QThread | None = None
        self._s_worker: _ScanarXMLsWorker | None = None
        self._a_thread: QThread | None = None
        self._a_worker: _AplicarCodigosWorker | None = None

        self.setWindowTitle("Atualizar Código Fornecedor via XML das NFe")
        self.setMinimumSize(1200, 640)
        if parent:
            self.setStyleSheet(parent.styleSheet())
        self.resize(1400, 750)

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        title = QLabel("Atualizar codigo_fornecedor via XML das NFe")
        title.setObjectName("sectionTitle")
        root.addWidget(title)

        info = QLabel(
            "<b>Regra de segurança:</b> Somente produtos onde "
            "<code>codigo_fornecedor = codigo_empresa</code> (cadastrados automaticamente) "
            "entram no escopo.<br>"
            "Matching: <b>chave NFe + quantidade + valor do item</b>. "
            "Ambíguo = mesmo código para produtos diferentes do mesmo fornecedor "
            "(ficam <b>desmarcados</b> por padrão)."
        )
        info.setWordWrap(True)
        root.addWidget(info)

        # Pasta de XMLs
        folder_row = QHBoxLayout()
        folder_row.addWidget(QLabel("Pasta de XMLs:"))
        self._folder_edit = QLineEdit()
        self._folder_edit.setPlaceholderText("Selecione a pasta com os arquivos .xml das NFe...")
        self._folder_edit.setReadOnly(True)
        folder_row.addWidget(self._folder_edit, 1)
        btn_browse = QPushButton("Procurar...")
        btn_browse.clicked.connect(self._browse_folder)
        folder_row.addWidget(btn_browse)
        self._btn_scan = QPushButton("Escanear e Propor")
        self._btn_scan.setObjectName("primaryButton")
        self._btn_scan.setEnabled(False)
        self._btn_scan.clicked.connect(self._run_scan)
        folder_row.addWidget(self._btn_scan)
        root.addLayout(folder_row)

        self._status_label = QLabel("")
        self._status_label.setObjectName("muted")
        root.addWidget(self._status_label)

        # Seleção rápida
        sel_row = QHBoxLayout()
        for label, fn in (
            ("Selecionar Todos",     lambda: self._set_all(True)),
            ("Desmarcar Ambíguos",   self._deselect_ambiguous),
            ("Desmarcar Todos",      lambda: self._set_all(False)),
        ):
            btn = QPushButton(label)
            btn.clicked.connect(fn)
            sel_row.addWidget(btn)
        sel_row.addStretch()
        self._count_label = QLabel("")
        sel_row.addWidget(self._count_label)
        root.addLayout(sel_row)

        # Tabela de preview
        self._table = QTableWidget()
        self._table.setColumnCount(len(self._PROPOSAL_HEADERS))
        self._table.setHorizontalHeaderLabels(self._PROPOSAL_HEADERS)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.verticalHeader().setVisible(False)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setAlternatingRowColors(False)
        self._table.setItemDelegate(_BrushDelegate(self._table))
        self._table.cellClicked.connect(self._on_cell_clicked)
        root.addWidget(self._table, 1)

        self._progress = QProgressBar()
        self._progress.hide()
        root.addWidget(self._progress)

        bottom = QHBoxLayout()
        self._btn_apply = QPushButton("Aplicar Atualizações Selecionadas")
        self._btn_apply.setObjectName("primaryButton")
        self._btn_apply.setEnabled(False)
        self._btn_apply.clicked.connect(self._apply)
        bottom.addWidget(self._btn_apply)
        bottom.addStretch()
        btn_fechar = QPushButton("Fechar")
        btn_fechar.clicked.connect(self.reject)
        bottom.addWidget(btn_fechar)
        root.addLayout(bottom)

    # ── Pasta ─────────────────────────────────────────────────────────────────

    def _browse_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(
            self, "Selecionar pasta com XMLs das NFe", ""
        )
        if folder:
            self._folder_edit.setText(folder)
            self._btn_scan.setEnabled(True)

    # ── Scan ──────────────────────────────────────────────────────────────────

    def _run_scan(self) -> None:
        folder = self._folder_edit.text().strip()
        if not folder:
            return
        self._btn_scan.setEnabled(False)
        self._btn_apply.setEnabled(False)
        self._progress.setRange(0, 100)
        self._progress.setValue(0)
        self._progress.show()
        self._status_label.setText("Escaneando XMLs...")

        self._s_thread = QThread(self)
        self._s_worker = _ScanarXMLsWorker(self._rows, self._empresa_id, self._repo, folder)
        self._s_worker.moveToThread(self._s_thread)
        self._s_thread.started.connect(self._s_worker.run)
        self._s_worker.progress.connect(self._on_scan_progress)
        self._s_worker.finished.connect(self._on_scan_done)
        self._s_worker.failed.connect(self._on_scan_failed)
        self._s_worker.finished.connect(self._s_thread.quit)
        self._s_worker.failed.connect(self._s_thread.quit)
        self._s_thread.finished.connect(self._s_worker.deleteLater)
        self._s_thread.finished.connect(self._s_thread.deleteLater)
        self._s_thread.finished.connect(lambda: setattr(self, "_s_thread", None))
        self._s_thread.finished.connect(lambda: setattr(self, "_s_worker", None))
        self._s_thread.start()

    def _on_scan_progress(self, current: int, total: int, msg: str) -> None:
        if total > 0:
            self._progress.setValue(int(current / total * 100))
        self._status_label.setText(msg)

    def _on_scan_done(self, proposals: list[dict]) -> None:
        self._progress.hide()
        self._btn_scan.setEnabled(True)
        self._proposals = proposals
        self._fill_table()
        n_uniq = sum(1 for p in proposals if p["confidence"] == "Único")
        n_amb  = sum(1 for p in proposals if p["confidence"] == "Ambíguo")
        if not proposals:
            self._status_label.setText(
                "Nenhuma proposta encontrada. Verifique se os XMLs correspondem às NFes do SPED "
                "e se há produtos com codigo_fornecedor = codigo_empresa."
            )
        else:
            self._status_label.setText(
                f"{len(proposals)} proposta(s) — {n_uniq} únicas (selecionadas) · "
                f"{n_amb} ambíguas (desmarcadas). Revise e clique em Aplicar."
            )
        self._update_count()

    def _on_scan_failed(self, msg: str) -> None:
        self._progress.hide()
        self._btn_scan.setEnabled(True)
        QMessageBox.critical(self, "Erro ao Escanear", f"Erro:\n{msg}")

    # ── Tabela de preview ─────────────────────────────────────────────────────

    def _fill_table(self) -> None:
        self._table.setUpdatesEnabled(False)
        self._table.setRowCount(len(self._proposals))
        for r, p in enumerate(self._proposals):
            is_unico = p["confidence"] == "Único"
            row_bg   = _ROW_EVEN if r % 2 == 0 else _ROW_ODD

            # Coluna 0: checkbox
            chk = QTableWidgetItem()
            chk.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable)
            chk.setCheckState(Qt.Checked if p["selected"] else Qt.Unchecked)
            chk.setBackground(row_bg)
            self._table.setItem(r, 0, chk)

            for c, field in enumerate(self._PROPOSAL_FIELDS[1:], start=1):
                val  = str(p.get(field) or "")
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                if c == 1:  # coluna Confiança
                    bg = self._BG_UNICO if is_unico else self._BG_AMBIG
                    fg = self._FG_UNICO if is_unico else self._FG_AMBIG
                    item.setBackground(bg)
                    item.setForeground(fg)
                    font = item.font(); font.setBold(True); item.setFont(font)
                elif c == 4:  # Cód. Novo (XML) — destaca em verde
                    item.setBackground(self._BG_UNICO if is_unico else self._BG_AMBIG)
                else:
                    item.setBackground(row_bg)
                self._table.setItem(r, c, item)

        self._table.resizeColumnsToContents()
        self._table.setUpdatesEnabled(True)

    def _on_cell_clicked(self, row: int, col: int) -> None:
        if col != 0 or row >= len(self._proposals):
            return
        item = self._table.item(row, 0)
        if item is None:
            return
        new_state = Qt.Unchecked if item.checkState() == Qt.Checked else Qt.Checked
        item.setCheckState(new_state)
        self._proposals[row]["selected"] = (new_state == Qt.Checked)
        self._update_count()

    def _set_all(self, state: bool) -> None:
        for r, p in enumerate(self._proposals):
            p["selected"] = state
            item = self._table.item(r, 0)
            if item:
                item.setCheckState(Qt.Checked if state else Qt.Unchecked)
        self._update_count()

    def _deselect_ambiguous(self) -> None:
        for r, p in enumerate(self._proposals):
            if p["confidence"] == "Ambíguo":
                p["selected"] = False
                item = self._table.item(r, 0)
                if item:
                    item.setCheckState(Qt.Unchecked)
        self._update_count()

    def _update_count(self) -> None:
        selected = sum(1 for p in self._proposals if p.get("selected"))
        self._count_label.setText(f"{selected}/{len(self._proposals)} selecionados")
        self._btn_apply.setEnabled(selected > 0)

    # ── Aplicar ───────────────────────────────────────────────────────────────

    def _apply(self) -> None:
        to_do = [p for p in self._proposals if p.get("selected")]
        if not to_do:
            return
        resp = QMessageBox.question(
            self, "Aplicar Atualizações",
            f"{len(to_do)} produto(s) terão o campo <b>codigo_fornecedor</b> atualizado.\n\n"
            "Essa operação só altera produtos onde codigo_fornecedor = codigo_empresa.\n"
            "Deseja continuar?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if resp != QMessageBox.Yes:
            return

        self._btn_apply.setEnabled(False)
        self._btn_scan.setEnabled(False)
        self._progress.setRange(0, 100)
        self._progress.setValue(0)
        self._progress.show()

        self._a_thread = QThread(self)
        self._a_worker = _AplicarCodigosWorker(to_do, self._repo)
        self._a_worker.moveToThread(self._a_thread)
        self._a_thread.started.connect(self._a_worker.run)
        self._a_worker.progress.connect(self._on_apply_progress)
        self._a_worker.finished.connect(self._on_apply_done)
        self._a_worker.failed.connect(self._on_apply_failed)
        self._a_worker.finished.connect(self._a_thread.quit)
        self._a_worker.failed.connect(self._a_thread.quit)
        self._a_thread.finished.connect(self._a_worker.deleteLater)
        self._a_thread.finished.connect(self._a_thread.deleteLater)
        self._a_thread.finished.connect(lambda: setattr(self, "_a_thread", None))
        self._a_thread.finished.connect(lambda: setattr(self, "_a_worker", None))
        self._a_thread.start()

    def _on_apply_progress(self, current: int, total: int, msg: str) -> None:
        if total > 0:
            self._progress.setValue(int(current / total * 100))
        self._status_label.setText(msg)

    def _on_apply_done(self, stats: dict) -> None:
        self._progress.hide()
        self._btn_scan.setEnabled(True)
        erros = stats.get("erros", [])
        n_pulados = stats.get("pulados", 0)
        txt = f"Atualizados: {stats['atualizados']}"
        if n_pulados:
            txt += f" | Pulados (código duplicado): {n_pulados}"
        txt += f" | Erros: {len(erros) - n_pulados}"
        if erros:
            txt += "\n\nDetalhes:\n" + "\n".join(f"  [{e['code']}] {e['erro']}" for e in erros[:10])
        QMessageBox.information(self, "Atualização Concluída", txt)
        applied_ids = {p["product_id"] for p in self._proposals if p.get("selected")}
        self._proposals = [p for p in self._proposals if p["product_id"] not in applied_ids]
        self._fill_table()
        self._update_count()
        self._status_label.setText(
            f"Atualizados: {stats['atualizados']}. "
            f"Propostas restantes: {len(self._proposals)}"
        )

    def _on_apply_failed(self, msg: str) -> None:
        self._progress.hide()
        self._btn_scan.setEnabled(True)
        self._btn_apply.setEnabled(True)
        QMessageBox.critical(self, "Erro ao Aplicar", f"Erro:\n{msg}")
